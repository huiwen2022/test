import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tksheet import Sheet
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

class CustomersTab:
    def __init__(self, parent, main_app):
        self.parent = parent
        self.main_app = main_app
        self.data = []
        self.columns = ["客戶編號","客戶名稱","聯絡人","電話","Email","地址","備註"]
        self.required_fields = ["客戶編號","客戶名稱","聯絡人"]
        self.dropdown_options = {}
        self.multiline_fields = ["地址","備註"]
        self.name_field = "客戶名稱"
        self.current_row = None
        self.setup_ui()

    # 共用工具與 UI 與 ProductsTab 相同（複製簡化）
    def _bind_mousewheel_to(self, canvas):
        def _on_wheel(event):
            if event.num in (4,5): delta = -1 if event.num == 4 else 1
            else: delta = -1 * (event.delta // 120)
            canvas.yview_scroll(delta, "units")
        canvas.bind("<Enter>", lambda e: (canvas.bind_all("<MouseWheel>", _on_wheel),
                                          canvas.bind_all("<Button-4>", _on_wheel),
                                          canvas.bind_all("<Button-5>", _on_wheel)))
        canvas.bind("<Leave>", lambda e: (canvas.unbind_all("<MouseWheel>"),
                                          canvas.unbind_all("<Button-4>"),
                                          canvas.unbind_all("<Button-5>")))
    def _setup_scrollable_edit_area(self, parent):
        container = ttk.Frame(parent); container.pack(fill=tk.BOTH, expand=True)
        canvas = tk.Canvas(container, highlightthickness=0)
        vbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vbar.set)
        inner = ttk.Frame(canvas)
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0), window=inner, anchor="nw")
        canvas.pack(side="left", fill="both", expand=True); vbar.pack(side="right", fill="y")
        self._bind_mousewheel_to(canvas); return inner
    def _make_input_widget(self, parent, col):
        if col in self.multiline_fields:
            return tk.Text(parent, height=3, wrap="word"), "text"
        return ttk.Entry(parent), "entry"
    def _set_widget_value(self,w,t,v):
        if t=="text": w.delete("1.0","end"); w.insert("1.0", v or ""); return
        w.delete(0, tk.END); w.insert(0, v or "")
    def _get_widget_value(self,w,t):
        return w.get("1.0","end-1c") if t=="text" else w.get()

    def _load_row_to_form(self, r):
        if r is None or not (0 <= r < len(self.data)): return
        vals = self.data[r]
        for i,col in enumerate(self.columns):
            w,t = self.inputs[col]; self._set_widget_value(w,t, vals[i] if i<len(vals) else "")
        self.current_row = r
    def _on_sheet_select(self,*_):
        sel = getattr(self.sheet,"get_currently_selected",lambda:None)(); r=getattr(sel,"row",None) if sel else None; self._load_row_to_form(r)

    def _validate(self, vals):
        for col in self.required_fields:
            i = self.columns.index(col)
            if not str(vals[i]).strip():
                messagebox.showerror("錯誤", f"{col} 為必填欄位"); return False
        return True

    def _update_row_heights_for_wrap(self):
        heights=[]; base=24
        for row in self.data:
            max_lines=1
            for col in self.multiline_fields:
                i=self.columns.index(col); txt=str(row[i]) if i < len(row) else ""
                lines=txt.count("\n")+1 if txt else 1
                max_lines=max(max_lines, lines)
            heights.append(min(base*max_lines, 24*6))
        try: self.sheet.set_row_heights(heights)
        except: pass

    def setup_ui(self):
        self.frame = ttk.Frame(self.parent); self.frame.pack(fill=tk.BOTH, expand=True)
        self.paned_window = ttk.PanedWindow(self.frame, orient=tk.VERTICAL); self.paned_window.pack(fill=tk.BOTH, expand=True)
        top = ttk.Frame(self.paned_window); self.paned_window.add(top, weight=3)
        toolbar = ttk.Frame(top); toolbar.pack(fill=tk.X, padx=5, pady=5)
        ttk.Button(toolbar, text="刪除選中", command=self.delete_row).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="清除全部", command=self.clear_all_data).pack(side=tk.LEFT, padx=2)
        ttk.Separator(toolbar, orient="vertical").pack(side=tk.LEFT, padx=10, fill=tk.Y)
        ttk.Button(toolbar, text="匯出Excel", command=self.export_to_excel).pack(side=tk.LEFT, padx=2)
        wrap = ttk.Frame(top); wrap.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.sheet = Sheet(wrap, headers=self.columns, data=self.data, table_wrap="w", header_wrap="w", column_width=120)
        self.sheet.enable_bindings("single_select","row_select","drag_select","select_all")
        self.sheet.grid(row=0,column=0,sticky="nsew"); wrap.grid_rowconfigure(0,weight=1); wrap.grid_columnconfigure(0,weight=1)
        getattr(self.sheet,"extra_bindings",lambda *a,**k: None)([("cell_select",self._on_sheet_select),("row_select",self._on_sheet_select)])
        try: self.sheet.bind("<ButtonRelease-1>", self._on_sheet_select)
        except: pass
        self.sheet.set_column_widths({i:w*7 for i,w in enumerate([12,20,14,16,26,30,24])})
        self.apply_sheet_styles()
        # 下方：可滾動編輯區

        bottom = ttk.Frame(self.paned_window); self.paned_window.add(bottom, weight=2)
        form_outer = self._setup_scrollable_edit_area(bottom)
        # 按鈕列在上方
        btns = ttk.Frame(form_outer); btns.grid(row=0, column=0, columnspan=4, sticky="w", pady=(2,6))
        ttk.Button(btns, text="新增" + ("" if "CustomersTab"!="OrdersTab" else "訂單"), command=self.add_row).pack(side=tk.LEFT, padx=4)
        ttk.Button(btns, text="修改" + ("" if "CustomersTab"!="OrdersTab" else "訂單"), command=self.update_row).pack(side=tk.LEFT, padx=4)
        ttk.Button(btns, text="清空表單", command=self.clear_form).pack(side=tk.LEFT, padx=4)

        # 兩欄並排的欄位（Label+輸入框）
        self.inputs = {}
        r = 1
        c = 0
        for col in self.columns:
            # 多行欄位改為獨占一列
            is_multiline = (col in getattr(self, "multiline_fields", []))
            if is_multiline:
                ttk.Label(form_outer, text=col + (" *" if col in self.required_fields else "")).grid(row=r, column=0, sticky="w", padx=5, pady=3)
                w, t = self._make_input_widget(form_outer, col)
                try:
                    w.configure(height=3)
                except Exception:
                    pass
                w.grid(row=r, column=1, columnspan=3, sticky="we", padx=5, pady=3)
                self.inputs[col] = (w, t)
                r += 1
                c = 0
                continue

            ttk.Label(form_outer, text=col + (" *" if col in self.required_fields else "")).grid(row=r, column=c*2, sticky="w", padx=5, pady=3)
            w, t = self._make_input_widget(form_outer, col)
            w.grid(row=r, column=c*2+1, sticky="we", padx=5, pady=3)
            self.inputs[col] = (w, t)
            c += 1
            if c == 2:
                c = 0
                r += 1

        # 欄位區域調整
        for i in range(4):
            form_outer.grid_columnconfigure(i, weight=1)
        ttk.Label(bottom, text="* 為必填欄位", foreground="red").pack(pady=(2,4))
    
    def get_selected_row(self):
        sel = getattr(self.sheet,"get_currently_selected", lambda: None)()
        return getattr(sel,"row",None) if sel else None
    def refresh_sheet(self):
        self.sheet.set_sheet_data(self.data, reset_highlights=True); self._update_row_heights_for_wrap(); self.apply_sheet_styles(); self.sheet.redraw()
    def add_row(self):
        vals=[self._get_widget_value(*self.inputs[c]) for c in self.columns]
        if not self._validate(vals): return
        self.data.append(vals); self.refresh_sheet(); self.clear_form()
    def update_row(self):
        r=self.get_selected_row()
        if r is None or not (0<=r<len(self.data)): messagebox.showwarning("警告","請先選取要修改的列"); return
        vals=[self._get_widget_value(*self.inputs[c]) for c in self.columns]
        if not self._validate(vals): return
        self.data[r]=vals; self.refresh_sheet(); self._load_row_to_form(r)
    def delete_row(self):
        r=self.get_selected_row()
        if r is None: messagebox.showwarning("警告","請先選取要刪除的列"); return
        del self.data[r]; self.refresh_sheet()
    def clear_all_data(self):
        if messagebox.askyesno("確認","確定清除全部資料？"): self.data.clear(); self.refresh_sheet()
    def clear_form(self):
        for c in self.columns:
            w,t = self.inputs[c]
            (w.delete("1.0","end") if t=="text" else w.delete(0, tk.END))
    def set_data(self, rows): self.data = rows or []; self.refresh_sheet()
    def get_data(self): return self.data
    def apply_sheet_styles(self):
        self.sheet.dehighlight_all(); rows=len(self.data); cols=len(self.columns)
        even=[r for r in range(rows) if (r+1)%2==0]; 
        if even: self.sheet.highlight_rows(even, bg="#E6F2FF")
        first5=list(range(min(5, cols))); others=list(range(5, cols))
        if first5: self.sheet.highlight_columns(first5, bg="#366092", fg="#FFFFFF", highlight_header=True)
        if others: self.sheet.highlight_columns(others, bg="#FFA500", fg="#000000", highlight_header=True)
    def export_to_excel(self):
        if not self.data: messagebox.showwarning("警告","沒有資料可以匯出"); return
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not filename: return
        wb=Workbook(); ws=wb.active; ws.title="客戶管理"
        BLUE=PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ORANGE=PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        EVEN=PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
        for i,t in enumerate(self.columns,1):
            c=ws.cell(row=1,column=i,value=t)
            if i<=5: c.fill=BLUE; c.font=Font(color="FFFFFF", bold=True)
            else: c.fill=ORANGE; c.font=Font(color="000000", bold=True)
            c.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r,row in enumerate(self.data,start=2):
            for c,v in enumerate(row,start=1):
                cell=ws.cell(row=r,column=c,value=v); cell.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
                if r%2==0: cell.fill=EVEN
        for i,w in enumerate([12,20,14,16,26,30,24], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes="B2"; wb.save(filename); messagebox.showinfo("成功", f"已匯出：{os.path.basename(filename)}")

    def load_data_from_list(self, rows):
        """舊版相容：從 list 載入資料"""
        self.set_data(rows or [])

    def load_sample_data(self):
        """舊版相容：載入範例資料（僅在無資料時）"""
        if not self.data:
            sample_data = [
            [
                        "C001",
                        "台北科技公司",
                        "王經理",
                        "02-12345678",
                        "manager@taipei-tech.com",
                        "台北市信義區松高路1號",
                        "長期合作客戶"
            ],
            [
                        "C002",
                        "高雄貿易商行",
                        "李先生",
                        "07-87654321",
                        "li@kaohsiung-trade.com",
                        "高雄市前鎮區中山三路100號",
                        "每季固定下單"
            ],
            [
                        "C003",
                        "台中製造廠",
                        "陳廠長",
                        "04-11223344",
                        "chen@taichung.com",
                        "台中市西屯區文心路200號",
                        "客製化需求多"
            ]
]
            self.set_data(sample_data)

