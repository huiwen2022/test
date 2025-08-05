import tkinter as tk
from tkinter import ttk, messagebox
import os
import sys
import json
# 取得 libs 資料夾的絕對路徑並加到 sys.path
sys.path.append(os.path.join(os.path.dirname(__file__), "libs"))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# 添加當前目錄到路徑，確保能導入其他模組
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from tab1_products import ProductsTab
from tab2_customers import CustomersTab
from tab3_orders import OrdersTab

class MainApplication:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel資料管理系統")
        self.root.geometry("1200x700")
        
        # JSON資料檔案路徑
        self.data_file = "data.json"
        
        # 創建主框架
        self.main_frame = ttk.Frame(root)
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 創建頁籤控制器
        self.notebook = ttk.Notebook(self.main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        # 初始化頁籤
        self.init_tabs()
        
        # 添加菜單欄
        self.create_menu()
        
        # 載入資料
        self.load_data()
    
    def init_tabs(self):
        """初始化所有頁籤"""
        try:
            # 產品管理頁籤
            self.products_tab = ProductsTab(self.notebook, self)
            self.notebook.add(self.products_tab.frame, text="產品管理")
            
            # 客戶管理頁籤
            self.customers_tab = CustomersTab(self.notebook, self)
            self.notebook.add(self.customers_tab.frame, text="客戶管理")
            
            # 訂單管理頁籤
            self.orders_tab = OrdersTab(self.notebook, self)
            self.notebook.add(self.orders_tab.frame, text="訂單管理")
            
        except Exception as e:
            messagebox.showerror("錯誤", f"初始化頁籤時發生錯誤: {str(e)}")
    
    def create_menu(self):
        """創建菜單欄"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # 檔案菜單
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="檔案", menu=file_menu)
        file_menu.add_command(label="匯出全部Excel", command=self.export_all_to_single_file)
        file_menu.add_separator()
        file_menu.add_command(label="清除全部資料", command=self.clear_all_data)
        file_menu.add_separator()
        file_menu.add_command(label="退出", command=self.on_closing)
        
        # 說明菜單
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="說明", menu=help_menu)
        help_menu.add_command(label="關於", command=self.show_about)
    
    def save_data(self):
        """儲存所有資料到JSON檔案"""
        try:
            data = {
                "products": self.products_tab.data,
                "customers": self.customers_tab.data,
                "orders": self.orders_tab.data
            }
            with open(self.data_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"儲存資料錯誤: {str(e)}")
    
    def load_data(self):
        """從JSON檔案載入資料"""
        try:
            if os.path.exists(self.data_file):
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                # 載入各頁籤資料
                if "products" in data:
                    self.products_tab.load_data_from_list(data["products"])
                if "customers" in data:
                    self.customers_tab.load_data_from_list(data["customers"])
                if "orders" in data:
                    self.orders_tab.load_data_from_list(data["orders"])
            else:
                # 如果沒有資料檔案，載入範例資料
                self.products_tab.load_sample_data()
                self.customers_tab.load_sample_data()
                self.orders_tab.load_sample_data()
        except Exception as e:
            print(f"載入資料錯誤: {str(e)}")
            # 如果載入失敗，載入範例資料
            self.products_tab.load_sample_data()
            self.customers_tab.load_sample_data()
            self.orders_tab.load_sample_data()
    
    def clear_all_data(self):
        """清除所有資料"""
        if messagebox.askyesno("確認", "確定要清除所有資料嗎？此操作無法復原！"):
            # 清除各頁籤資料
            self.products_tab.clear_all_data()
            self.customers_tab.clear_all_data()
            self.orders_tab.clear_all_data()
            
            # 刪除JSON檔案
            try:
                if os.path.exists(self.data_file):
                    os.remove(self.data_file)
                messagebox.showinfo("成功", "已清除所有資料！")
            except Exception as e:
                messagebox.showerror("錯誤", f"清除資料時發生錯誤: {str(e)}")
    
    def export_all_to_single_file(self):
        """匯出所有頁籤到單一Excel檔案"""
        try:
            from tkinter import filedialog
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="儲存整合Excel檔案"
            )
            if not filename:
                return
            wb = Workbook()
            BLUE = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ORANGE = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            EVEN = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
            GREEN = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
            GRAY  = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")

            tabs = [
                ("產品管理", self.products_tab.columns, self.products_tab.get_data()),
                ("客戶管理", self.customers_tab.columns, self.customers_tab.get_data()),
                ("訂單管理", self.orders_tab.columns, self.orders_tab.get_data()),
            ]

            for sheet_name, columns, data in tabs:
                ws = wb.create_sheet(title=sheet_name) if wb.worksheets else wb.active
                ws.title = sheet_name

                # 標題列（前5欄藍底白字，其餘橘底黑字）
                for col_num, column_title in enumerate(columns, 1):
                    cell = ws.cell(row=1, column=col_num, value=column_title)
                    if col_num <= 5:
                        cell.fill = BLUE; cell.font = Font(color="FFFFFF", bold=True)
                    else:
                        cell.fill = ORANGE; cell.font = Font(color="000000", bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                if data:
                    for row_num, row_data in enumerate(data, start=2):
                        row_even = (row_num % 2 == 0)
                        for col_num, value in enumerate(row_data, start=1):
                            cell = ws.cell(row=row_num, column=col_num, value=value)
                            # 訂單管理數字欄位靠右
                            if sheet_name == "訂單管理" and col_num in [4, 5, 6]:
                                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                            else:
                                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                            if row_even:
                                cell.fill = EVEN

                        # 狀態規則（訂單管理：第8欄）
                        if sheet_name == "訂單管理":
                            st = str(ws.cell(row=row_num, column=8).value or "").strip()
                            if st == "已完成":
                                ws.cell(row=row_num, column=8).fill = GREEN
                            elif st == "已取消":
                                for c in range(1, len(columns)+1):
                                    ws.cell(row=row_num, column=c).fill = GRAY
                else:
                    cell = ws.cell(row=2, column=1, value="目前沒有資料")
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.font = Font(italic=True, color="999999")

                # 凍結首列首欄
                ws.freeze_panes = "B2"

                # 固定欄寬
                # from openpyxl.utils import get_column_letter
                # if sheet_name == "產品管理":
                #     widths = [12,18,12,10,10,14,24]
                # elif sheet_name == "客戶管理":
                #     widths = [12,20,14,16,26,30,24]
                # else:  # 訂單管理
                #     widths = [12,18,20,10,12,14,14,12]
                # for i,w in enumerate(widths, start=1):
                #     ws.column_dimensions[get_column_letter(i)].width = w

            # 移除預設空白表（若存在且無資料）
            if "Sheet" in [s.title for s in wb.worksheets] and len(wb.worksheets) > 3:
                for s in wb.worksheets:
                    if s.title == "Sheet":
                        wb.remove(s); break

            wb.save(filename)
            messagebox.showinfo("成功", f"已匯出：{os.path.basename(filename)}")
        except Exception as e:
            messagebox.showerror("錯誤", f"匯出Excel時發生錯誤：{str(e)}")

    
    def export_all(self):
        """匯出所有頁籤的資料（舊版本，保留兼容性）"""
        self.export_all_to_single_file()
    
    def on_closing(self):
        """程式關閉時的處理"""
        self.save_data()
        self.root.quit()
    
    def show_about(self):
        """顯示關於資訊"""
        messagebox.showinfo("關於", "Excel資料管理系統 v2.0\n\n功能包含：\n- 多頁籤資料管理\n- Excel匯出功能\n- 自訂樣式格式\n- JSON資料儲存\n- 必填欄位驗證\n- 下拉選單支援")

def main():
    # 確保必要的模組存在
    required_files = ['tab1_products.py', 'tab2_customers.py', 'tab3_orders.py']
    missing_files = []
    
    for file in required_files:
        if not os.path.exists(file):
            missing_files.append(file)
    
    if missing_files:
        print(f"警告：缺少以下檔案: {', '.join(missing_files)}")
        print("請確保所有頁籤檔案都在同一目錄下")
    
    root = tk.Tk()
    app = MainApplication(root)
    
    # 設置關閉事件
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    
    root.mainloop()

if __name__ == "__main__":
    main()