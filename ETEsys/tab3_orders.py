import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tksheet import Sheet
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

class OrdersTab:
    def __init__(self, parent, main_app):
        self.parent = parent; self.main_app = main_app; self.data = []
        self.columns = ["訂單編號","客戶名稱","產品名稱","數量","單價","總金額","訂單日期","狀態"]
        self.required_fields = ["訂單編號","客戶名稱","產品名稱","數量","單價"]
        self.dropdown_options = {"狀態": ["待處理","處理中","待出貨","已出貨","已完成","已取消"]}
        self.multiline_fields = []
        self.name_field = "產品名稱"
        self.current_row = None
        self.setup_ui()

    # （與前兩個類似的工具函式）
    def _bind_mousewheel_to(self, canvas):
        def _on_wheel(event):
            if event.num in (4,5): delta = -1 if event.num == 4 else 1
            else: delta = -1 * (event.delta // 120)
            canvas.yview_scroll(delta, "units")
        canvas.bind("<Enter>", lambda e: (canvas.bind_all("<MouseWheel>", _on_wheel),
                                          canvas.bind_all("<Button-4>", _on_wheel),
                                          canvas.bind_all("<Button-5>", _on_wheel)))
        canvas.bind("<Leave>", lambda e: (canvas.unbind_all("<MouseWheel>"),
                                          canvas.unbind_all("<Button-4>"),
                                          canvas.unbind_all("<Button-5>")))
    def _setup_scrollable_edit_area(self, parent):
        container = ttk.Frame(parent); container.pack(fill=tk.BOTH, expand=True)
        canvas = tk.Canvas(container, highlightthickness=0)
        vbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vbar.set)
        inner = ttk.Frame(canvas)
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0,0), window=inner, anchor="nw")
        canvas.pack(side="left", fill="both", expand=True); vbar.pack(side="right", fill="y")
        self._bind_mousewheel_to(canvas); return inner
    def _make_input_widget(self, parent, col):
        if col in self.dropdown_options: return ttk.Combobox(parent, values=self.dropdown_options[col], state="readonly"), "combo"
        return ttk.Entry(parent), "entry"
    def _set_widget_value(self,w,t,v):
        if t=="combo": w.set("" if v is None else str(v)); return
        w.delete(0, tk.END); w.insert(0, "" if v is None else str(v))
    def _get_widget_value(self,w,t):
        return w.get()

    def _load_row_to_form(self, r):
        if r is None or not (0 <= r < len(self.data)): return
        vals=self.data[r]
        for i,col in enumerate(self.columns):
            w,t=self.inputs[col]; self._set_widget_value(w,t, vals[i] if i<len(vals) else "")
        self.current_row=r
    def _on_sheet_select(self,*_):
        sel = getattr(self.sheet,"get_currently_selected",lambda:None)(); r=getattr(sel,"row",None) if sel else None; self._load_row_to_form(r)

    def _validate(self, vals):
        for col in self.required_fields:
            i=self.columns.index(col)
            if not str(vals[i]).strip():
                messagebox.showerror("錯誤", f"{col} 為必填欄位"); return False
        # 數字欄位
        for col in ["數量","單價","總金額"]:
            i=self.columns.index(col)
            s=str(vals[i]).strip()
            if s:
                try: float(s)
                except: messagebox.showerror("錯誤", f"{col} 必須是數字"); return False
        # 日期
        i=self.columns.index("訂單日期")
        s=str(vals[i]).strip()
        if s:
            try: datetime.strptime(s, "%Y-%m-%d")
            except: messagebox.showerror("錯誤","訂單日期格式需為 YYYY-MM-DD"); return False
        return True

    def _update_row_heights_for_wrap(self):
        try: self.sheet.set_row_heights([24]*len(self.data))
        except: pass

    def setup_ui(self):
        self.frame = ttk.Frame(self.parent); self.frame.pack(fill=tk.BOTH, expand=True)
        self.paned_window = ttk.PanedWindow(self.frame, orient=tk.VERTICAL); self.paned_window.pack(fill=tk.BOTH, expand=True)
        top = ttk.Frame(self.paned_window); self.paned_window.add(top, weight=3)
        toolbar = ttk.Frame(top); toolbar.pack(fill=tk.X, padx=5, pady=5)
        ttk.Button(toolbar, text="刪除選中", command=self.delete_row).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="複製選中", command=self.copy_row).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="清除全部", command=self.clear_all_data).pack(side=tk.LEFT, padx=2)
        ttk.Separator(toolbar, orient="vertical").pack(side=tk.LEFT, padx=10, fill=tk.Y)
        ttk.Button(toolbar, text="匯出Excel", command=self.export_to_excel).pack(side=tk.LEFT, padx=2)
        wrap = ttk.Frame(top); wrap.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.sheet = Sheet(wrap, headers=self.columns, data=self.data, table_wrap="w", header_wrap="w", column_width=120)
        self.sheet.enable_bindings("single_select","row_select","drag_select","select_all")
        self.sheet.grid(row=0,column=0,sticky="nsew"); wrap.grid_rowconfigure(0,weight=1); wrap.grid_columnconfigure(0,weight=1)
        getattr(self.sheet,"extra_bindings",lambda *a,**k: None)([("cell_select",self._on_sheet_select),("row_select",self._on_sheet_select)])
        try: self.sheet.bind("<ButtonRelease-1>", self._on_sheet_select)
        except: pass
        self.sheet.set_column_widths({i:w*7 for i,w in enumerate([12,18,20,10,12,14,14,12])})
        self.apply_sheet_styles()
        bottom = ttk.Frame(self.paned_window); self.paned_window.add(bottom, weight=2)
        form = self._setup_scrollable_edit_area(bottom)
        self.inputs = {}
        for col in self.columns:
            row = ttk.Frame(form); row.pack(fill=tk.X, pady=4)
            ttk.Label(row, text=col + (" *" if col in self.required_fields else "")).pack(side=tk.LEFT, padx=5)
            w,t = self._make_input_widget(row,col); w.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5); self.inputs[col]=(w,t)
        btns = ttk.Frame(bottom); btns.pack(pady=6)
        ttk.Button(btns, text="新增訂單", command=self.add_row).pack(side=tk.LEFT, padx=5)
        ttk.Button(btns, text="修改訂單", command=self.update_row).pack(side=tk.LEFT, padx=5)
        ttk.Button(btns, text="清空表單", command=self.clear_form).pack(side=tk.LEFT, padx=5)

    def get_selected_row(self):
        sel=getattr(self.sheet,"get_currently_selected",lambda:None)(); return getattr(sel,"row",None) if sel else None
    def refresh_sheet(self):
        self.sheet.set_sheet_data(self.data, reset_highlights=True); self._update_row_heights_for_wrap(); self.apply_sheet_styles(); self.sheet.redraw()
    def add_row(self):
        vals=[self._get_widget_value(*self.inputs[c]) for c in self.columns]
        # 自動算總金額
        try: iqty=self.columns.index("數量"); ip=self.columns.index("單價"); it=self.columns.index("總金額"); qty=float(vals[iqty] or 0); price=float(vals[ip] or 0); vals[it]=qty*price
        except: pass
        if not self._validate(vals): return
        self.data.append(vals); self.refresh_sheet(); self.clear_form()
    def update_row(self):
        r=self.get_selected_row()
        if r is None or not (0<=r<len(self.data)): messagebox.showwarning("警告","請先選取要修改的列"); return
        vals=[self._get_widget_value(*self.inputs[c]) for c in self.columns]
        try: iqty=self.columns.index("數量"); ip=self.columns.index("單價"); it=self.columns.index("總金額"); qty=float(vals[iqty] or 0); price=float(vals[ip] or 0); vals[it]=qty*price
        except: pass
        if not self._validate(vals): return
        self.data[r]=vals; self.refresh_sheet(); self._load_row_to_form(r)
    def copy_row(self):
        r=self.get_selected_row()
        if r is None or not (0<=r<len(self.data)): messagebox.showwarning("警告","請先選取要複製的列"); return
        new_row=list(self.data[r])
        # 名稱做區別
        try:
            idx=self.columns.index(self.name_field)
            base=str(new_row[idx]); cand=f"{base} - 複本"; existing={str(row[idx]) for row in self.data if len(row)>idx}
            if cand in existing:
                n=2
                while f"{base} - 複本{n}" in existing: n+=1
                cand=f"{base} - 複本{n}"
            new_row[idx]=cand
        except: pass
        self.data.append(new_row); self.refresh_sheet()
    def delete_row(self):
        r=self.get_selected_row()
        if r is None: messagebox.showwarning("警告","請先選取要刪除的列"); return
        del self.data[r]; self.refresh_sheet()
    def clear_all_data(self):
        if messagebox.askyesno("確認","確定清除全部資料？"): self.data.clear(); self.refresh_sheet()
    def clear_form(self):
        for c in self.columns:
            w,t=self.inputs[c]
            (w.set("") if t=="combo" else w.delete(0, tk.END))
    def set_data(self, rows): self.data=rows or []; self.refresh_sheet()
    def get_data(self): return self.data
    def apply_sheet_styles(self):
        self.sheet.dehighlight_all()
        rows=len(self.data); cols=len(self.columns)
        even=[r for r in range(rows) if (r+1)%2==0]
        if even: self.sheet.highlight_rows(even, bg="#E6F2FF")
        first5=list(range(min(5, cols))); others=list(range(5, cols))
        if first5: self.sheet.highlight_columns(first5, bg="#366092", fg="#FFFFFF", highlight_header=True)
        if others: self.sheet.highlight_columns(others, bg="#FFA500", fg="#000000", highlight_header=True)
        # 狀態規則
        for r,row in enumerate(self.data):
            st=str(row[7]).strip() if len(row)>7 else ""
            if st=="已完成": self.sheet.highlight_cells(row=r, column=7, bg="#D5F5E3")
            elif st=="已取消": self.sheet.highlight_rows([r], bg="#EEEEEE")
        for c in [3,4,5]: self.sheet.align_columns([c], align="e")
    def export_to_excel(self):
        if not self.data: messagebox.showwarning("警告","沒有資料可以匯出"); return
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not filename: return
        wb=Workbook(); ws=wb.active; ws.title="訂單管理"
        BLUE=PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ORANGE=PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        EVEN=PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
        GREEN=PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
        GRAY=PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
        for i,t in enumerate(self.columns,1):
            c=ws.cell(row=1,column=i,value=t)
            if i<=5: c.fill=BLUE; c.font=Font(color="FFFFFF", bold=True)
            else: c.fill=ORANGE; c.font=Font(color="000000", bold=True)
            c.alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r,row in enumerate(self.data,start=2):
            even = (r%2==0)
            for c,v in enumerate(row,start=1):
                cell=ws.cell(row=r,column=c,value=v)
                cell.alignment=Alignment(horizontal=("right" if c in [4,5,6] else "center"), vertical="center", wrap_text=True)
                if even: cell.fill=EVEN
            st=str(ws.cell(row=r,column=8).value or "").strip()
            if st=="已完成": ws.cell(row=r,column=8).fill=GREEN
            elif st=="已取消":
                for c in range(1,len(self.columns)+1): ws.cell(row=r,column=c).fill=GRAY
        for i,w in enumerate([12,18,20,10,12,14,14,12], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes="B2"; wb.save(filename); messagebox.showinfo("成功", f"已匯出：{os.path.basename(filename)}")

    def load_data_from_list(self, rows):
        """舊版相容：從 list 載入資料"""
        self.set_data(rows or [])

    def load_sample_data(self):
        """舊版相容：載入範例資料（僅在無資料時）"""
        if not self.data:
            sample_data = [
            [
                        "O001",
                        "台北科技公司",
                        "筆記型電腦",
                        "2",
                        "25000",
                        "50000",
                        "2024-01-15",
                        "已完成"
            ],
            [
                        "O002",
                        "高雄貿易商行",
                        "無線滑鼠",
                        "10",
                        "800",
                        "8000",
                        "2024-01-16",
                        "處理中"
            ],
            [
                        "O003",
                        "台中製造廠",
                        "辦公椅",
                        "5",
                        "3500",
                        "17500",
                        "2024-01-17",
                        "待出貨"
            ]
]
            self.set_data(sample_data)

