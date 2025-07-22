#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel處理模組
提供Excel檔案的進階讀寫功能
"""

import os
import sys
from datetime import datetime, date

# 設置套件路徑
def setup_environment():
    base_path = os.path.dirname(os.path.abspath(__file__))
    libs_path = os.path.join(base_path, 'libs')
    
    if os.path.exists(libs_path) and libs_path not in sys.path:
        sys.path.insert(0, libs_path)

setup_environment()

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import ColorScaleRule
    print("✅ Excel處理模組載入成功")
except ImportError as e:
    print(f"❌ Excel處理模組載入失敗: {e}")
    raise


class ExcelHandler:
    """Excel處理類別"""
    
    def __init__(self):
        self.workbook = None
        self.current_sheet = None
        
        # 預設樣式定義
        self.styles = {
            'header': {
                'font': Font(bold=True, color="FFFFFF", size=12),
                'fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
                'alignment': Alignment(horizontal="center", vertical="center"),
                'border': Border(left=Side(style='medium'), right=Side(style='medium'),
                               top=Side(style='medium'), bottom=Side(style='medium'))
            },
            'data': {
                'font': Font(size=10),
                'alignment': Alignment(horizontal="left", vertical="center"),
                'border': Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))
            },
            'highlight_green': {
                'fill': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            },
            'highlight_red': {
                'fill': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            },
            'highlight_yellow': {
                'fill': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            }
        }
    
    def create_workbook(self):
        """創建新的工作簿"""
        self.workbook = openpyxl.Workbook()
        # 移除預設工作表
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        return self.workbook
    
    def load_workbook(self, file_path):
        """載入現有工作簿"""
        try:
            self.workbook = openpyxl.load_workbook(file_path)
            return self.workbook
        except Exception as e:
            raise Exception(f"載入Excel檔案失敗: {str(e)}")
    
    def create_sheet(self, sheet_name, headers, data=None):
        """創建工作表"""
        if not self.workbook:
            self.create_workbook()
        
        # 創建工作表
        worksheet = self.workbook.create_sheet(title=sheet_name)
        self.current_sheet = worksheet
        
        # 寫入標題行
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            self.apply_style(cell, self.styles['header'])
        
        # 寫入資料
        if data:
            self.write_data_to_sheet(worksheet, data, headers)
        
        # 應用格式設定
        self.format_sheet(worksheet, len(headers), len(data) + 1 if data else 1)
        
        return worksheet
    
    def write_data_to_sheet(self, worksheet, data, headers):
        """將資料寫入工作表"""
        for row_num, record in enumerate(data, 2):  # 從第2行開始
            for col, header in enumerate(headers, 1):
                # 根據標題找對應的資料鍵值
                field_key = self.get_field_key_from_header(header)
                value = record.get(field_key, "")
                
                # 處理日期格式
                if isinstance(value, (date, datetime)):
                    value = value.strftime("%Y-%m-%d")
                
                cell = worksheet.cell(row=row_num, column=col, value=value)
                self.apply_style(cell, self.styles['data'])
                
                # 根據內容應用條件格式
                self.apply_conditional_formatting(cell, header, value)
    
    def apply_style(self, cell, style_dict):
        """應用樣式到儲存格"""
        if 'font' in style_dict:
            cell.font = style_dict['font']
        if 'fill' in style_dict:
            cell.fill = style_dict['fill']
        if 'alignment' in style_dict:
            cell.alignment = style_dict['alignment']
        if 'border' in style_dict:
            cell.border = style_dict['border']
    
    def apply_conditional_formatting(self, cell, header, value):
        """應用條件格式"""
        if header in ["狀態", "Status"]:
            if value == "已核准" or value == "Approved":
                self.apply_style(cell, self.styles['highlight_green'])
            elif value == "已拒絕" or value == "Rejected":
                self.apply_style(cell, self.styles['highlight_red'])
            elif value == "待審核" or value == "Pending":
                self.apply_style(cell, self.styles['highlight_yellow'])
        
        elif header in ["考績", "Performance", "年度總評"]:
            if value == "優" or value == "Excellent":
                self.apply_style(cell, self.styles['highlight_green'])
            elif value == "差" or value == "Poor":
                self.apply_style(cell, self.styles['highlight_red'])
    
    def format_sheet(self, worksheet, num_cols, num_rows):
        """格式化工作表"""
        # 自動調整欄寬
        for col in range(1, num_cols + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            
            for row in range(1, num_rows + 1):
                cell = worksheet.cell(row=row, column=col)
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # 設定欄寬（最小10，最大50）
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 凍結首列
        worksheet.freeze_panes = 'A2'
        
        # 加入自動篩選
        if num_rows > 1:
            worksheet.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{num_rows}"
    
    def add_data_validation(self, worksheet, cell_range, validation_list):
        """加入資料驗證（下拉選單）"""
        dv = DataValidation(type="list", formula1=f'"{",".join(validation_list)}"', allow_blank=True)
        dv.error = '請從下拉選單中選擇'
        dv.errorTitle = '輸入錯誤'
        dv.prompt = '請選擇一個選項'
        dv.promptTitle = '選擇選項'
        
        worksheet.add_data_validation(dv)
        dv.add(cell_range)
    
    def create_summary_sheet(self, data_dict):
        """創建摘要工作表"""
        if not self.workbook:
            self.create_workbook()
        
        summary_sheet = self.workbook.create_sheet(title="資料摘要", index=0)
        
        # 建立摘要資訊
        summary_data = [
            ["項目", "數量", "最後更新"],
            ["基本資料", 1 if data_dict.get('basic_info') else 0, datetime.now().strftime("%Y-%m-%d %H:%M")],
            ["考績記錄", len(data_dict.get('performance_records', [])), datetime.now().strftime("%Y-%m-%d %H:%M")],
            ["出勤記錄", len(data_dict.get('attendance_records', [])), datetime.now().strftime("%Y-%m-%d %H:%M")],
            ["請假申請", len(data_dict.get('leave_requests', [])), datetime.now().strftime("%Y-%m-%d %H:%M")],
            ["加班申請", len(data_dict.get('overtime_requests', [])), datetime.now().strftime("%Y-%m-%d %H:%M")]
        ]
        
        # 寫入摘要資料
        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = summary_sheet.cell(row=row_num, column=col_num, value=value)
                
                if row_num == 1:  # 標題行
                    self.apply_style(cell, self.styles['header'])
                else:
                    self.apply_style(cell, self.styles['data'])
        
        # 格式化摘要表
        self.format_sheet(summary_sheet, 3, len(summary_data))
        
        # 加入圖表統計
        self.add_summary_statistics(summary_sheet, data_dict)
        
        return summary_sheet
    
    def add_summary_statistics(self, worksheet, data_dict):
        """加入統計資訊"""
        # 考績統計
        if data_dict.get('performance_records'):
            performance_stats = self.calculate_performance_stats(data_dict['performance_records'])
            
            # 寫入考績統計
            start_row = 8
            worksheet.cell(row=start_row, column=1, value="考績統計")
            self.apply_style(worksheet.cell(row=start_row, column=1), self.styles['header'])
            
            for i, (grade, count) in enumerate(performance_stats.items(), 1):
                worksheet.cell(row=start_row + i, column=1, value=grade)
                worksheet.cell(row=start_row + i, column=2, value=count)
                self.apply_style(worksheet.cell(row=start_row + i, column=1), self.styles['data'])
                self.apply_style(worksheet.cell(row=start_row + i, column=2), self.styles['data'])
        
        # 出勤統計
        if data_dict.get('attendance_records'):
            attendance_stats = self.calculate_attendance_stats(data_dict['attendance_records'])
            
            start_row = 15
            worksheet.cell(row=start_row, column=1, value="出勤統計")
            self.apply_style(worksheet.cell(row=start_row, column=1), self.styles['header'])
            
            for i, (status, count) in enumerate(attendance_stats.items(), 1):
                worksheet.cell(row=start_row + i, column=1, value=status)
                worksheet.cell(row=start_row + i, column=2, value=count)
                self.apply_style(worksheet.cell(row=start_row + i, column=1), self.styles['data'])
                self.apply_style(worksheet.cell(row=start_row + i, column=2), self.styles['data'])
    
    def calculate_performance_stats(self, performance_data):
        """計算考績統計"""
        stats = {}
        for record in performance_data:
            rating = record.get('annual_rating', '未知')
            stats[rating] = stats.get(rating, 0) + 1
        return stats
    
    def calculate_attendance_stats(self, attendance_data):
        """計算出勤統計"""
        stats = {}
        for record in attendance_data:
            status = record.get('status', '未知')
            stats[status] = stats.get(status, 0) + 1
        return stats
    
    def get_field_key_from_header(self, header):
        """根據表頭獲取欄位鍵值"""
        header_mapping = {
            # 基本資料
            "員工編號": "employee_id",
            "姓名": "name",
            "身分證字號": "id_number",
            "性別": "gender",
            "出生日期": "birth_date",
            "聯絡電話": "phone",
            "電子郵件": "email",
            "緊急聯絡人": "emergency_contact",
            "緊急聯絡人電話": "emergency_phone",
            "戶籍地址": "address",
            "通訊地址": "mailing_address",
            "部門": "department",
            "職位": "position",
            "職級": "job_level",
            "到職日期": "hire_date",
            "直屬主管": "supervisor",
            "工作地點": "work_location",
            "僱用類型": "employment_type",
            "薪資等級": "salary_grade",
            
            # 考績資料
            "年度": "year",
            "上半年考績": "first_half",
            "下半年考績": "second_half",
            "年度總評": "annual_rating",
            "備註": "remarks",
            
            # 出勤資料
            "日期": "date",
            "上班時間": "start_time",
            "下班時間": "end_time",
            "工作時數": "hours",
            "狀態": "status",
            
            # 請假資料
            "請假類型": "leave_type",
            "開始日期": "start_date",
            "結束日期": "end_date",
            "請假天數": "days",
            "申請日期": "apply_date",
            "請假事由": "reason",
            
            # 加班資料
            "加班日期": "overtime_date",
            "加班時數": "hours",
            "加班類型": "overtime_type",
            "加班事由": "reason"
        }
        
        return header_mapping.get(header, header.lower().replace(" ", "_").replace("*", ""))
    
    def read_excel_data(self, file_path, sheet_name=None):
        """讀取Excel資料"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    raise Exception(f"工作表 '{sheet_name}' 不存在")
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            
            # 讀取資料
            data = []
            headers = []
            
            for row_num, row in enumerate(worksheet.iter_rows(values_only=True), 1):
                if row_num == 1:
                    headers = [str(cell).strip() if cell is not None else f"Column_{i}" 
                              for i, cell in enumerate(row)]
                else:
                    if any(cell is not None for cell in row):
                        row_data = {}
                        for i, cell in enumerate(row):
                            if i < len(headers):
                                field_key = self.get_field_key_from_header(headers[i])
                                value = cell if cell is not None else ""
                                
                                # 處理日期格式
                                if isinstance(value, datetime):
                                    value = value.strftime("%Y-%m-%d")
                                elif isinstance(value, date):
                                    value = value.strftime("%Y-%m-%d")
                                
                                row_data[field_key] = str(value).strip() if value else ""
                        
                        if any(row_data.values()):  # 如果行中有任何非空值
                            data.append(row_data)
            
            return {
                'headers': headers,
                'data': data,
                'sheet_name': worksheet.title
            }
            
        except Exception as e:
            raise Exception(f"讀取Excel資料時發生錯誤: {str(e)}")
    
    def save_workbook(self, file_path):
        """儲存工作簿"""
        if not self.workbook:
            raise Exception("沒有工作簿可以儲存")
        
        try:
            self.workbook.save(file_path)
            return True
        except Exception as e:
            raise Exception(f"儲存Excel檔案時發生錯誤: {str(e)}")
    
    def create_template(self, template_type="employee"):
        """創建範本檔案"""
        self.create_workbook()
        
        if template_type == "employee":
            self.create_employee_template()
        elif template_type == "attendance":
            self.create_attendance_template()
        elif template_type == "performance":
            self.create_performance_template()
        
        return self.workbook
    
    def create_employee_template(self):
        """創建員工資料範本"""
        # 基本資料範本
        basic_headers = [
            "員工編號", "姓名", "身分證字號", "性別", "出生日期",
            "聯絡電話", "電子郵件", "部門", "職位", "到職日期"
        ]
        
        basic_sheet = self.create_sheet("基本資料範本", basic_headers)
        
        # 加入範例資料
        example_data = [{
            "employee_id": "EMP001",
            "name": "張三",
            "id_number": "A123456789",
            "gender": "男",
            "birth_date": "1990-01-01",
            "phone": "0912345678",
            "email": "zhang@example.com",
            "department": "技術部",
            "position": "工程師",
            "hire_date": "2020-01-15"
        }]
        
        self.write_data_to_sheet(basic_sheet, example_data, basic_headers)
        
        # 加入資料驗證
        self.add_data_validation(basic_sheet, "D2:D1000", ["男", "女"])  # 性別
        self.add_data_validation(basic_sheet, "H2:H1000", ["人事部", "財務部", "業務部", "技術部", "行政部"])  # 部門
    
    def create_attendance_template(self):
        """創建出勤記錄範本"""
        headers = ["日期", "員工編號", "姓名", "上班時間", "下班時間", "狀態", "備註"]
        sheet = self.create_sheet("出勤記錄範本", headers)
        
        # 加入資料驗證
        self.add_data_validation(sheet, "F2:F1000", ["正常", "遲到", "早退", "曠職", "請假"])
    
    def create_performance_template(self):
        """創建考績記錄範本"""
        headers = ["員工編號", "姓名", "年度", "上半年考績", "下半年考績", "年度總評", "備註"]
        sheet = self.create_sheet("考績記錄範本", headers)
        
        # 加入資料驗證
        rating_options = ["優", "良", "可", "差"]
        self.add_data_validation(sheet, "D2:D1000", rating_options)  # 上半年考績
        self.add_data_validation(sheet, "E2:E1000", rating_options)  # 下半年考績
        self.add_data_validation(sheet, "F2:F1000", rating_options)  # 年度總評


# 工具函數
def create_employee_excel(data_dict, output_path):
    """快速創建員工Excel檔案"""
    handler = ExcelHandler()
    handler.create_workbook()
    
    # 創建摘要頁
    handler.create_summary_sheet(data_dict)
    
    # 創建各個資料頁
    if data_dict.get('basic_info'):
        basic_headers = ["欄位", "內容"]
        basic_data = [{"field": k, "content": v} for k, v in data_dict['basic_info'].items()]
        handler.create_sheet("基本資料", basic_headers, basic_data)
    
    if data_dict.get('performance_records'):
        perf_headers = ["年度", "上半年考績", "下半年考績", "年度總評", "備註"]
        handler.create_sheet("考績記錄", perf_headers, data_dict['performance_records'])
    
    if data_dict.get('attendance_records'):
        att_headers = ["日期", "上班時間", "下班時間", "工作時數", "狀態", "備註"]
        handler.create_sheet("出勤記錄", att_headers, data_dict['attendance_records'])
    
    if data_dict.get('leave_requests'):
        leave_headers = ["請假類型", "開始日期", "結束日期", "請假天數", "申請日期", "狀態", "請假事由"]
        handler.create_sheet("請假記錄", leave_headers, data_dict['leave_requests'])
    
    if data_dict.get('overtime_requests'):
        ot_headers = ["加班日期", "開始時間", "結束時間", "加班時數", "加班類型", "申請日期", "狀態", "加班事由"]
        handler.create_sheet("加班記錄", ot_headers, data_dict['overtime_requests'])
    
    # 儲存檔案
    handler.save_workbook(output_path)
    return True


def create_excel_template(template_type="employee", output_path="template.xlsx"):
    """創建Excel範本檔案"""
    handler = ExcelHandler()
    handler.create_template(template_type)
    handler.save_workbook(output_path)
    return True


# 測試用主程式
if __name__ == "__main__":
    print("Excel處理模組測試")
    
    # 測試創建範本
    try:
        create_excel_template("employee", "員工資料範本.xlsx")
        print("✅ 員工資料範本創建成功")
    except Exception as e:
        print(f"❌ 範本創建失敗: {e}")
    
    # 測試讀取資料
    try:
        handler = ExcelHandler()
        # 這裡可以測試讀取功能，需要有測試檔案
        print("✅ Excel處理模組測試完成")
    except Exception as e:
        print(f"❌ 測試失敗: {e}")