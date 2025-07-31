#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
主程式 - Excel 報表生成器
讀取多個 CSV 檔案並生成格式化的 Excel 報表
"""
import os
import sys

# 加入 libs 路徑（確保是 openpyxl 模組所在的那層）
base_path = os.path.dirname(os.path.abspath(__file__))
libs_path = os.path.join(base_path, 'libs')

if libs_path not in sys.path:
    sys.path.insert(0, libs_path)
    import openpyxl

from sheet1_generator import Sheet1Generator
from sheet2_generator import Sheet2Generator

# 設定標準輸出編碼為 UTF-8
if sys.platform.startswith('win'):
    # Windows 系統設定
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.detach())
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.detach())
    # 設定控制台編碼
    os.system('chcp 65001 >nul 2>&1')

def main():
    # CSV 檔案路徑
    csv_files = {
        'sheet1': 'data1.csv',  # 第一個 sheet 的 CSV 檔案
        'sheet2': 'data2.csv',  # 第二個 sheet 的 CSV 檔案
    }
    
    # 輸出的 Excel 檔案名稱
    output_file = 'formatted_report.xlsx'
    
    # 檢查 CSV 檔案是否存在
    missing_files = []
    for sheet_name, file_path in csv_files.items():
        if not os.path.exists(file_path):
            missing_files.append(file_path)
    
    if missing_files:
        print("錯誤：以下 CSV 檔案不存在：")
        for file in missing_files:
            print(f"  - {file}")
        print("請確認檔案路徑是否正確。")
        return
    
    try:
        # 建立新的 Excel 工作簿
        workbook = openpyxl.Workbook()
        # 刪除預設工作表
        workbook.remove(workbook.active)
        
        # 建立第一個工作表
        print(f"正在處理 {csv_files['sheet1']} -> Sheet1")
        sheet1_generator = Sheet1Generator()
        sheet1_generator.create_sheet(workbook, csv_files['sheet1'], 'Sheet1')
        
        # 建立第二個工作表
        print(f"正在處理 {csv_files['sheet2']} -> Sheet2")
        sheet2_generator = Sheet2Generator()
        sheet2_generator.create_sheet(workbook, csv_files['sheet2'], 'Sheet2')
        
        # 儲存 Excel 檔案
        workbook.save(output_file)
        print(f"✅ Excel 檔案已成功生成：{output_file}")
        
    except Exception as e:
        print(f"❌ 處理過程中發生錯誤：{str(e)}")

if __name__ == "__main__":
    main()