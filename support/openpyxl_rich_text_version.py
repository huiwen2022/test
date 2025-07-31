#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
使用 openpyxl 的 Rich Text 功能（修正版）
可以在同一個儲存格內使用多種字體顏色
"""

import csv
import sys
import os

# 加入 libs 路徑（確保是 openpyxl 模組所在的那層）
base_path = os.path.dirname(os.path.abspath(__file__))
libs_path = os.path.join(base_path, 'libs')

if libs_path not in sys.path:
    sys.path.insert(0, libs_path)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.rich_text import TextBlock, CellRichText
    from openpyxl.cell.text import InlineFont  # 修正：使用 InlineFont

# 設定標準輸出編碼為 UTF-8
if sys.platform.startswith('win'):
    import codecs
    if hasattr(sys.stdout, 'detach'):
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.detach())
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.detach())

class OpenpyxlRichTextGenerator:
    def __init__(self):
        """初始化工作簿"""
        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)
        
        # 定義 InlineFont 樣式（Rich Text 專用）
        self.inline_fonts = {
            'red_font': InlineFont(color='FF0000', b=True),      # b=True 代表 bold
            'green_font': InlineFont(color='008000', b=True),    # b=True 代表 bold
            'normal_font': InlineFont(color='000000'),
        }
        
        # 定義一般字體樣式（標題用）
        self.fonts = {
            'white_font': Font(color='FFFFFF', bold=True),
            'black_font': Font(color='000000', bold=True),
        }
        
        # 定義填充樣式
        self.fills = {
            'blue_bg': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
            'orange_bg': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),
            'yellow_bg': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
        }
        
        # 定義對齊樣式（包含文字換行）
        self.alignments = {
            'center_wrap': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'left_wrap': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'general_wrap': Alignment(horizontal='center', vertical='center', wrap_text=True)  # 改為垂直水平居中
        }
    
    def read_csv_data(self, csv_file):
        """讀取 CSV 檔案數據"""
        data = []
        encodings = ['utf-8-sig', 'utf-8', 'big5', 'cp950', 'gbk', 'gb2312']
        
        for encoding in encodings:
            try:
                with open(csv_file, 'r', encoding=encoding, newline='') as file:
                    sample = file.read(1024)
                    file.seek(0)
                    
                    if '�' not in sample:
                        sniffer = csv.Sniffer()
                        try:
                            delimiter = sniffer.sniff(sample).delimiter
                        except:
                            delimiter = ','
                        
                        file.seek(0)
                        reader = csv.reader(file, delimiter=delimiter)
                        for row in reader:
                            data.append(row)
                        print(f"✅ 使用 {encoding} 編碼成功讀取 {csv_file}")
                        return data
            except UnicodeDecodeError:
                continue
            except Exception as e:
                print(f"嘗試編碼 {encoding} 時發生錯誤: {e}")
                continue
        
        raise ValueError(f"無法讀取 CSV 檔案 {csv_file}，請檢查檔案編碼格式")
    
    def create_rich_text_instruction(self):
        """建立 Rich Text 說明文字"""
        # 建立 Rich Text 物件
        rich_text = CellRichText()
        
        # 添加不同顏色的文字片段（使用 InlineFont）
        rich_text.append(TextBlock(self.inline_fonts['red_font'], "紅色"))
        rich_text.append(TextBlock(self.inline_fonts['normal_font'], "為注意事項，"))
        rich_text.append(TextBlock(self.inline_fonts['green_font'], "綠色"))
        rich_text.append(TextBlock(self.inline_fonts['normal_font'], "為通過"))
        
        return rich_text
    
    def create_sheet1(self, csv_file):
        """建立 Sheet1"""
        data = self.read_csv_data(csv_file)
        if not data:
            raise ValueError(f"CSV 檔案 {csv_file} 是空的或無法讀取")
        
        worksheet = self.workbook.create_sheet(title='Sheet1')
        
        # 設置 Rich Text 說明文字 (B1)
        rich_text = self.create_rich_text_instruction()
        worksheet['B1'].value = rich_text
        
        # 合併 B1 到最後一欄
        num_cols = len(data[0])
        last_col_letter = get_column_letter(num_cols + 1)
        merge_range = f"B1:{last_col_letter}1"
        worksheet.merge_cells(merge_range)
        
        # 從 B2 開始填入標題
        headers = data[0]
        for col_idx, header in enumerate(headers):
            cell = worksheet.cell(row=2, column=col_idx + 2)
            cell.value = header
            
            # 前5欄藍色，其餘橘色
            if col_idx < 5:
                cell.fill = self.fills['blue_bg']
                cell.font = self.fonts['white_font']
            else:
                cell.fill = self.fills['orange_bg']
                cell.font = self.fonts['black_font']
            
            # 設置對齊和文字換行
            cell.alignment = self.alignments['center_wrap']
        
        # 填入數據 (從 B3 開始)
        for row_idx, row_data in enumerate(data[1:], start=3):
            for col_idx, cell_value in enumerate(row_data):
                cell = worksheet.cell(row=row_idx, column=col_idx + 2)
                cell.value = cell_value
                
                # 第四欄設置黃色背景
                if col_idx == 3:  # 第四欄 (E欄)
                    cell.fill = self.fills['yellow_bg']
                
                # 所有儲存格都支援文字換行
                cell.alignment = self.alignments['general_wrap']
                
                # 所有儲存格都支援文字換行
                cell.alignment = self.alignments['general_wrap']
        
        # 設置篩選器
        if len(data) > 1:
            filter_range = f"B2:{last_col_letter}{len(data) + 1}"
            worksheet.auto_filter.ref = filter_range
        
        # 凍結標題列和第一欄
        worksheet.freeze_panes = 'C3'
        
        # 設置欄寬
        for col_idx in range(1, num_cols + 2):
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = 15
        
        print(f"✅ Sheet1 已創建完成 (前5欄藍色，後7欄橘色，凍結標題列+第一欄)")
    
    def create_sheet2(self, csv_file):
        """建立 Sheet2"""
        data = self.read_csv_data(csv_file)
        if not data:
            raise ValueError(f"CSV 檔案 {csv_file} 是空的或無法讀取")
        
        worksheet = self.workbook.create_sheet(title='Sheet2')
        
        # 設置 Rich Text 說明文字 (B1)
        rich_text = self.create_rich_text_instruction()
        worksheet['B1'].value = rich_text
        
        # 合併 B1 到最後一欄
        num_cols = len(data[0])
        last_col_letter = get_column_letter(num_cols + 1)
        merge_range = f"B1:{last_col_letter}1"
        worksheet.merge_cells(merge_range)
        
        # 從 B2 開始填入標題
        headers = data[0]
        for col_idx, header in enumerate(headers):
            cell = worksheet.cell(row=2, column=col_idx + 2)
            cell.value = header
            
            # 前8欄藍色，其餘橘色
            if col_idx < 8:
                cell.fill = self.fills['blue_bg']
                cell.font = self.fonts['white_font']
            else:
                cell.fill = self.fills['orange_bg']
                cell.font = self.fonts['black_font']
            
            # 設置對齊和文字換行
            cell.alignment = self.alignments['center_wrap']
        
        # 填入數據 (從 B3 開始)
        for row_idx, row_data in enumerate(data[1:], start=3):
            for col_idx, cell_value in enumerate(row_data):
                cell = worksheet.cell(row=row_idx, column=col_idx + 2)
                cell.value = cell_value
                
                # 第四欄設置黃色背景
                if col_idx == 3:  # 第四欄 (E欄)
                    cell.fill = self.fills['yellow_bg']
        
        # 設置篩選器
        if len(data) > 1:
            filter_range = f"B2:{last_col_letter}{len(data) + 1}"
            worksheet.auto_filter.ref = filter_range
        
        # 只凍結標題列
        worksheet.freeze_panes = 'B3'
        
        # 設置欄寬
        for col_idx in range(1, num_cols + 2):
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = 15
        
        print(f"✅ Sheet2 已創建完成 (前8欄藍色，後10欄橘色，只凍結標題列)")
    
    def save(self, filename):
        """儲存檔案"""
        self.workbook.save(filename)

def main():
    """主程式"""
    csv_files = {
        'sheet1': 'data1.csv',
        'sheet2': 'data2.csv',
    }
    
    output_file = 'formatted_report_richtext.xlsx'
    
    # 檢查檔案
    missing_files = []
    for sheet_name, file_path in csv_files.items():
        if not os.path.exists(file_path):
            missing_files.append(file_path)
    
    if missing_files:
        print("錯誤：以下 CSV 檔案不存在：")
        for file in missing_files:
            print(f"  - {file}")
        print("請確認檔案路徑是否正確。")
        return
    
    try:
        # 建立 Excel 生成器
        generator = OpenpyxlRichTextGenerator()
        
        # 建立工作表
        print(f"正在處理 {csv_files['sheet1']} -> Sheet1")
        generator.create_sheet1(csv_files['sheet1'])
        
        print(f"正在處理 {csv_files['sheet2']} -> Sheet2")
        generator.create_sheet2(csv_files['sheet2'])
        
        # 儲存檔案
        generator.save(output_file)
        print(f"✅ Excel 檔案已成功生成：{output_file}")
        
    except Exception as e:
        print(f"❌ 處理過程中發生錯誤：{str(e)}")

if __name__ == "__main__":
    main()