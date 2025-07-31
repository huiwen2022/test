#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sheet2 生成器
前8欄藍色、後10欄橘色，只凍結標題列
"""

import csv
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class Sheet2Generator:
    def __init__(self):
        """初始化 Sheet2 生成器"""
        # 定義顏色樣式
        self.colors = {
            'red_font': Font(color='FF0000', bold=True),      # 紅色字體
            'green_font': Font(color='008000', bold=True),    # 綠色字體
            'blue_bg': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),    # 藍色背景
            'orange_bg': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),  # 橘色背景
            'yellow_bg': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'), # 亮黃色背景
            'white_font': Font(color='FFFFFF', bold=True),    # 白色字體
            'black_font': Font(color='000000', bold=True),    # 黑色字體
        }
        
        # Sheet2 的設定
        self.blue_cols = 8      # 前8欄藍色
        self.orange_cols = 10   # 後10欄橘色
    
    def read_csv_data(self, csv_file):
        """讀取 CSV 檔案數據"""
        data = []
        # 嘗試多種編碼格式
        encodings = ['utf-8-sig', 'utf-8', 'big5', 'cp950', 'gbk', 'gb2312']
        
        for encoding in encodings:
            try:
                with open(csv_file, 'r', encoding=encoding, newline='') as file:
                    # 自動偵測分隔符號
                    sample = file.read(1024)
                    file.seek(0)
                    
                    # 如果讀取成功且沒有亂碼，使用這個編碼
                    if '�' not in sample:  # 檢查有沒有亂碼字符
                        sniffer = csv.Sniffer()
                        try:
                            delimiter = sniffer.sniff(sample).delimiter
                        except:
                            delimiter = ','  # 預設使用逗號
                        
                        file.seek(0)
                        reader = csv.reader(file, delimiter=delimiter)
                        for row in reader:
                            data.append(row)
                        print(f"✅ 使用 {encoding} 編碼成功讀取 {csv_file}")
                        return data
            except UnicodeDecodeError:
                continue
            except Exception as e:
                print(f"嘗試編碼 {encoding} 時發生錯誤: {e}")
                continue
        
        # 如果所有編碼都失敗，拋出錯誤
        raise ValueError(f"無法讀取 CSV 檔案 {csv_file}，請檢查檔案編碼格式")
    
    def create_sheet(self, workbook, csv_file, sheet_name):
        """建立 Sheet2"""
        # 讀取 CSV 數據
        data = self.read_csv_data(csv_file)
        if not data:
            raise ValueError(f"CSV 檔案 {csv_file} 是空的或無法讀取")
        
        # 建立工作表
        worksheet = workbook.create_sheet(title=sheet_name)
        
        # 設置說明文字 (B1)
        self._set_instruction_text(worksheet, len(data[0]) if data else 0)
        
        # 從 B2 開始填入數據
        self._fill_data(worksheet, data)
        
        # 設置格式
        self._format_worksheet(worksheet, len(data[0]) if data else 0, len(data))
        
        # 只凍結標題列
        worksheet.freeze_panes = 'B3'  # 只凍結 B2 以上
        
        print(f"✅ Sheet2 已創建完成 (前{self.blue_cols}欄藍色，後{self.orange_cols}欄橘色，只凍結標題列)")
    
    def _set_instruction_text(self, worksheet, num_cols):
        """設置 B1 的說明文字"""
        # 在 B1 設置純文字
        worksheet['B1'] = "紅色為注意事項，綠色為通過"
        
        # 在 C1 和 D1 分別設置紅色和綠色的文字說明
        worksheet['C1'] = "紅色"
        worksheet['C1'].font = self.colors['red_font']
        worksheet['D1'] = "綠色"
        worksheet['D1'].font = self.colors['green_font']
        
        # 合併 B1 到最後一欄
        if num_cols > 0:
            last_col_letter = get_column_letter(num_cols + 1)  # +1 因為從 B 欄開始
            merge_range = f"B1:{last_col_letter}1"
            worksheet.merge_cells(merge_range)
    
    def _fill_data(self, worksheet, data):
        """從 B2 開始填入數據"""
        headers = data[0]
        
        # 填入標題 (B2 行)
        for col_idx, header in enumerate(headers):
            cell = worksheet.cell(row=2, column=col_idx + 2)  # 從 B 欄開始
            cell.value = header
        
        # 填入數據 (從 B3 開始)
        for row_idx, row_data in enumerate(data[1:], start=3):
            for col_idx, cell_value in enumerate(row_data):
                cell = worksheet.cell(row=row_idx, column=col_idx + 2)
                cell.value = cell_value
    
    def _format_worksheet(self, worksheet, num_cols, num_rows):
        """設置 Sheet2 的格式"""
        if num_cols == 0:
            return
        
        # 設置標題行格式 (B2 行)
        for col_idx in range(num_cols):
            col_letter = get_column_letter(col_idx + 2)  # 從 B 欄開始
            cell = worksheet[f'{col_letter}2']
            
            # 前8欄：藍色背景 + 白色字體
            if col_idx < self.blue_cols:
                cell.fill = self.colors['blue_bg']
                cell.font = self.colors['white_font']
            # 其餘欄位：橘色背景 + 黑色字體
            else:
                cell.fill = self.colors['orange_bg']
                cell.font = self.colors['black_font']
            
            # 設置對齊
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 設置第四欄 (E欄) 整欄為亮黃色背景
        fourth_col_letter = get_column_letter(5)  # 第四欄是 E 欄
        for row in range(2, num_rows + 2):  # 從標題行開始到最後一行
            cell = worksheet[f'{fourth_col_letter}{row}']
            cell.fill = self.colors['yellow_bg']
        
        # 設置篩選器 (自動篩選)
        if num_rows > 1:  # 確保有數據行
            last_col_letter = get_column_letter(num_cols + 1)
            filter_range = f"B2:{last_col_letter}{num_rows + 1}"
            worksheet.auto_filter.ref = filter_range
        
        # 自動調整欄寬
        for col_idx in range(1, num_cols + 2):  # 從 A 到最後一欄
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = 15