#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
員工表單系統 - 主程式
支援多員工管理、Excel匯入/匯出、多頁籤功能
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import sys
from datetime import datetime, date
import json

# 設置套件路徑
def setup_environment():
    """設置程式運行環境"""
    base_path = os.path.dirname(os.path.abspath(__file__))
    libs_path = os.path.join(base_path, 'libs')
    
    if os.path.exists(libs_path) and libs_path not in sys.path:
        sys.path.insert(0, libs_path)

setup_environment()

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    print("✅ openpyxl 載入成功")
except ImportError as e:
    messagebox.showerror("錯誤", f"無法載入 openpyxl: {e}")
    sys.exit(1)

class EmployeeFormSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("員工表單系統 v2.0")
        self.root.geometry("1400x900")
        
        # 資料儲存 - 改為管理多個員工
        self.employees_data = {}  # {employee_id: {basic_info: {}, performance_records: [], attendance_records: []}}
        self.current_employee_id = None
        
        # 建立GUI
        self.create_widgets()
        
    def create_widgets(self):
        """建立主要界面"""
        # 主框架
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 標題
        title_label = ttk.Label(main_frame, text="員工表單系統 v2.0", font=("Arial", 16, "bold"))
        title_label.pack(pady=(0, 10))
        
        # 工具列
        toolbar_frame = ttk.Frame(main_frame)
        toolbar_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(toolbar_frame, text="📁 匯入Excel", command=self.import_excel).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(toolbar_frame, text="💾 匯出Excel", command=self.export_excel).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(toolbar_frame, text="🗑️ 清空資料", command=self.clear_all_data).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(toolbar_frame, text="💾 儲存資料", command=self.save_data).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(toolbar_frame, text="📂 載入資料", command=self.load_data).pack(side=tk.LEFT, padx=(0, 5))
        
        # 分頁控件
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        # 建立各個頁籤
        self.create_employee_management_tab()
        self.create_performance_management_tab()
        self.create_attendance_management_tab()
        
    def create_employee_management_tab(self):
        """建立員工管理頁籤"""
        emp_frame = ttk.Frame(self.notebook)
        self.notebook.add(emp_frame, text="👥 員工管理")
        
        # 分割為左右兩個區域
        paned = ttk.PanedWindow(emp_frame, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 左側：員工列表
        left_frame = ttk.Frame(paned)
        paned.add(left_frame, weight=1)
        
        # 員工列表標題
        ttk.Label(left_frame, text="員工列表", font=("Arial", 12, "bold")).pack(pady=(0, 5))
        
        # 員工列表
        list_frame = ttk.Frame(left_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        # 員工樹狀檢視
        emp_columns = ('員工編號', '姓名', '部門', '職位', '到職日期')
        self.employee_tree = ttk.Treeview(list_frame, columns=emp_columns, show='headings', height=15)
        
        for col in emp_columns:
            self.employee_tree.heading(col, text=col)
            self.employee_tree.column(col, width=100)
        
        emp_scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.employee_tree.yview)
        self.employee_tree.configure(yscrollcommand=emp_scrollbar.set)
        
        self.employee_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        emp_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 綁定選擇事件
        self.employee_tree.bind('<<TreeviewSelect>>', self.on_employee_select)
        
        # 員工操作按鈕
        emp_button_frame = ttk.Frame(left_frame)
        emp_button_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(emp_button_frame, text="➕ 新增員工", command=self.new_employee).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(emp_button_frame, text="🗑️ 刪除員工", command=self.delete_employee).pack(side=tk.LEFT, padx=(0, 5))
        
        # 右側：員工詳細資料
        right_frame = ttk.Frame(paned)
        paned.add(right_frame, weight=2)
        
        # 建立滾動區域
        canvas = tk.Canvas(right_frame)
        scrollbar = ttk.Scrollbar(right_frame, orient="vertical", command=canvas.yview)
        self.emp_detail_frame = ttk.Frame(canvas)
        
        self.emp_detail_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=self.emp_detail_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 詳細資料標題
        ttk.Label(self.emp_detail_frame, text="員工詳細資料", font=("Arial", 12, "bold")).pack(pady=(0, 10))
        
        # 基本資料欄位
        self.basic_fields = {}
        
        # 員工基本資訊
        info_frame = ttk.LabelFrame(self.emp_detail_frame, text="員工基本資訊", padding=10)
        info_frame.pack(fill=tk.X, padx=5, pady=5)
        
        basic_info_fields = [
            ("員工編號*", "employee_id", "entry", True),
            ("姓名*", "name", "entry", True),
            ("身分證字號*", "id_number", "entry", True),
            ("性別*", "gender", "combobox", True, ["男", "女"]),
            ("出生日期*", "birth_date", "entry", True),
            ("聯絡電話*", "phone", "entry", True),
            ("電子郵件", "email", "entry", False),
            ("緊急聯絡人", "emergency_contact", "entry", False),
            ("緊急聯絡人電話", "emergency_phone", "entry", False),
            ("戶籍地址*", "address", "text", True),
            ("通訊地址", "mailing_address", "text", False),
        ]
        
        row = 0
        for field_info in basic_info_fields:
            field_name, field_key = field_info[0], field_info[1]
            field_type, required = field_info[2], field_info[3]
            
            ttk.Label(info_frame, text=field_name).grid(row=row, column=0, sticky=tk.W, padx=5, pady=2)
            
            if field_type == "entry":
                widget = ttk.Entry(info_frame, width=40)
            elif field_type == "combobox":
                widget = ttk.Combobox(info_frame, width=37, values=field_info[4] if len(field_info) > 4 else [])
                widget.state(['readonly'])
            elif field_type == "text":
                widget = tk.Text(info_frame, width=30, height=3)
            
            widget.grid(row=row, column=1, sticky=tk.W, padx=5, pady=2)
            self.basic_fields[field_key] = {'widget': widget, 'required': required}
            
            row += 1
        
        # 工作資訊
        work_frame = ttk.LabelFrame(self.emp_detail_frame, text="工作資訊", padding=10)
        work_frame.pack(fill=tk.X, padx=5, pady=5)
        
        work_info_fields = [
            ("部門*", "department", "combobox", True, ["人事部", "財務部", "業務部", "技術部", "行政部"]),
            ("職位*", "position", "combobox", True, ["經理", "副理", "主任", "專員", "助理"]),
            ("職級*", "job_level", "combobox", True, ["1級", "2級", "3級", "4級", "5級"]),
            ("到職日期*", "hire_date", "entry", True),
            ("直屬主管", "supervisor", "entry", False),
            ("工作地點*", "work_location", "combobox", True, ["台北", "台中", "高雄", "新竹"]),
            ("僱用類型*", "employment_type", "combobox", True, ["正職", "約聘", "派遣", "兼職"]),
            ("薪資等級", "salary_grade", "combobox", False, ["A級", "B級", "C級", "D級"]),
        ]
        
        row = 0
        for field_info in work_info_fields:
            field_name, field_key = field_info[0], field_info[1]
            field_type, required = field_info[2], field_info[3]
            
            ttk.Label(work_frame, text=field_name).grid(row=row, column=0, sticky=tk.W, padx=5, pady=2)
            
            if field_type == "entry":
                widget = ttk.Entry(work_frame, width=40)
            elif field_type == "combobox":
                widget = ttk.Combobox(work_frame, width=37, values=field_info[4] if len(field_info) > 4 else [])
                widget.state(['readonly'])
            
            widget.grid(row=row, column=1, sticky=tk.W, padx=5, pady=2)
            self.basic_fields[field_key] = {'widget': widget, 'required': required}
            
            row += 1
        
        # 按鈕
        button_frame = ttk.Frame(self.emp_detail_frame)
        button_frame.pack(fill=tk.X, padx=5, pady=10)
        
        ttk.Button(button_frame, text="💾 儲存員工資料", command=self.save_employee_info).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="🗑️ 清空表單", command=self.clear_employee_form).pack(side=tk.LEFT, padx=5)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
    def create_performance_management_tab(self):
        """建立考績管理頁籤"""
        perf_frame = ttk.Frame(self.notebook)
        self.notebook.add(perf_frame, text="📊 考績管理")
        
        # 上方：員工選擇
        selection_frame = ttk.LabelFrame(perf_frame, text="選擇員工", padding=5)
        selection_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(selection_frame, text="員工：").pack(side=tk.LEFT, padx=5)
        self.perf_employee_var = tk.StringVar()
        self.perf_employee_combo = ttk.Combobox(selection_frame, textvariable=self.perf_employee_var, 
                                              width=30, state="readonly")
        self.perf_employee_combo.pack(side=tk.LEFT, padx=5)
        self.perf_employee_combo.bind('<<ComboboxSelected>>', self.on_perf_employee_select)
        
        ttk.Button(selection_frame, text="🔄 重新整理", command=self.refresh_employee_combos).pack(side=tk.LEFT, padx=10)
        
        # 中間：考績記錄列表
        history_frame = ttk.LabelFrame(perf_frame, text="考績記錄", padding=5)
        history_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        perf_columns = ('年度', '上半年考績', '下半年考績', '年度總評', '備註')
        self.perf_tree = ttk.Treeview(history_frame, columns=perf_columns, show='headings', height=10)
        
        for col in perf_columns:
            self.perf_tree.heading(col, text=col)
            self.perf_tree.column(col, width=120)
        
        perf_scrollbar = ttk.Scrollbar(history_frame, orient=tk.VERTICAL, command=self.perf_tree.yview)
        self.perf_tree.configure(yscrollcommand=perf_scrollbar.set)
        
        self.perf_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        perf_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 下方：新增/編輯考績
        add_perf_frame = ttk.LabelFrame(perf_frame, text="新增/編輯考績記錄", padding=10)
        add_perf_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.perf_fields = {}
        
        perf_input_fields = [
            ("年度*", "year", "combobox", [str(year) for year in range(2020, 2030)]),
            ("上半年考績", "first_half", "combobox", ["優", "良", "可", "差"]),
            ("下半年考績", "second_half", "combobox", ["優", "良", "可", "差"]),
            ("年度總評*", "annual_rating", "combobox", ["優", "良", "可", "差"]),
            ("備註", "remarks", "text", None),
        ]
        
        row = 0
        for field_name, field_key, field_type, values in perf_input_fields:
            ttk.Label(add_perf_frame, text=field_name).grid(row=row, column=0, sticky=tk.W, padx=5, pady=2)
            
            if field_type == "combobox":
                widget = ttk.Combobox(add_perf_frame, width=37, values=values)
                if field_key in ['first_half', 'second_half', 'annual_rating']:
                    widget.state(['readonly'])
            elif field_type == "text":
                widget = tk.Text(add_perf_frame, width=30, height=3)
            
            widget.grid(row=row, column=1, sticky=tk.W, padx=5, pady=2)
            self.perf_fields[field_key] = widget
            
            row += 1
        
        perf_button_frame = ttk.Frame(add_perf_frame)
        perf_button_frame.grid(row=row, column=0, columnspan=2, pady=10)
        
        ttk.Button(perf_button_frame, text="➕ 新增考績", command=self.add_performance).pack(side=tk.LEFT, padx=5)
        ttk.Button(perf_button_frame, text="✏️ 編輯考績", command=self.edit_performance).pack(side=tk.LEFT, padx=5)
        ttk.Button(perf_button_frame, text="🗑️ 刪除考績", command=self.delete_performance).pack(side=tk.LEFT, padx=5)
        
    def create_attendance_management_tab(self):
        """建立出勤管理頁籤"""
        att_frame = ttk.Frame(self.notebook)
        self.notebook.add(att_frame, text="⏰ 出勤管理")
        
        # 分成兩個子頁籤
        att_notebook = ttk.Notebook(att_frame)
        att_notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 請假管理頁籤
        self.create_leave_management_tab(att_notebook)
        
        # 加班管理頁籤
        self.create_overtime_management_tab(att_notebook)
        
    def create_leave_management_tab(self, parent):
        """建立請假管理頁籤"""
        leave_frame = ttk.Frame(parent)
        parent.add(leave_frame, text="🏖️ 請假管理")
        
        # 上方：篩選區域
        filter_frame = ttk.LabelFrame(leave_frame, text="篩選條件", padding=5)
        filter_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # 員工篩選
        ttk.Label(filter_frame, text="員工：").grid(row=0, column=0, padx=5, pady=2)
        self.leave_employee_var = tk.StringVar()
        self.leave_employee_combo = ttk.Combobox(filter_frame, textvariable=self.leave_employee_var, 
                                               width=20, state="readonly")
        self.leave_employee_combo.grid(row=0, column=1, padx=5, pady=2)
        
        # 狀態篩選
        ttk.Label(filter_frame, text="狀態：").grid(row=0, column=2, padx=5, pady=2)
        self.leave_status_var = tk.StringVar()
        leave_status_combo = ttk.Combobox(filter_frame, textvariable=self.leave_status_var,
                                        values=["全部", "待審核", "已核准", "已拒絕"], width=15, state="readonly")
        leave_status_combo.grid(row=0, column=3, padx=5, pady=2)
        leave_status_combo.set("全部")
        
        ttk.Button(filter_frame, text="🔍 篩選", command=self.filter_leave_records).grid(row=0, column=4, padx=10, pady=2)
        ttk.Button(filter_frame, text="🔄 重新整理", command=self.refresh_leave_records).grid(row=0, column=5, padx=5, pady=2)
        
        # 中間：請假記錄列表
        leave_list_frame = ttk.LabelFrame(leave_frame, text="請假記錄", padding=5)
        leave_list_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        leave_columns = ('員工編號', '員工姓名', '請假類型', '開始日期', '結束日期', '請假天數', '申請日期', '狀態', '備註')
        self.leave_tree = ttk.Treeview(leave_list_frame, columns=leave_columns, show='headings', height=10)
        
        for col in leave_columns:
            self.leave_tree.heading(col, text=col)
            self.leave_tree.column(col, width=90)
        
        leave_scrollbar = ttk.Scrollbar(leave_list_frame, orient=tk.VERTICAL, command=self.leave_tree.yview)
        self.leave_tree.configure(yscrollcommand=leave_scrollbar.set)
        
        self.leave_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        leave_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 下方：新增請假申請
        add_leave_frame = ttk.LabelFrame(leave_frame, text="新增請假申請", padding=10)
        add_leave_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.leave_fields = {}
        
        # 第一行
        row1_frame = ttk.Frame(add_leave_frame)
        row1_frame.pack(fill=tk.X, pady=2)
        
        ttk.Label(row1_frame, text="員工*").grid(row=0, column=0, padx=5, pady=2, sticky=tk.W)
        self.leave_emp_var = tk.StringVar()
        self.leave_fields['employee'] = ttk.Combobox(row1_frame, textvariable=self.leave_emp_var, 
                                                   width=15, state="readonly")
        self.leave_fields['employee'].grid(row=0, column=1, padx=5, pady=2)
        
        ttk.Label(row1_frame, text="請假類型*").grid(row=0, column=2, padx=5, pady=2, sticky=tk.W)
        self.leave_fields['leave_type'] = ttk.Combobox(row1_frame, width=15, state="readonly",
                                                     values=["年假", "病假", "事假", "婚假", "喪假", "產假", "陪產假"])
        self.leave_fields['leave_type'].grid(row=0, column=3, padx=5, pady=2)
        
        ttk.Label(row1_frame, text="開始日期*").grid(row=0, column=4, padx=5, pady=2, sticky=tk.W)
        self.leave_fields['start_date'] = ttk.Entry(row1_frame, width=12)
        self.leave_fields['start_date'].grid(row=0, column=5, padx=5, pady=2)
        
        # 第二行
        row2_frame = ttk.Frame(add_leave_frame)
        row2_frame.pack(fill=tk.X, pady=2)
        
        ttk.Label(row2_frame, text="結束日期*").grid(row=0, column=0, padx=5, pady=2, sticky=tk.W)
        self.leave_fields['end_date'] = ttk.Entry(row2_frame, width=12)
        self.leave_fields['end_date'].grid(row=0, column=1, padx=5, pady=2)
        
        ttk.Label(row2_frame, text="請假天數*").grid(row=0, column=2, padx=5, pady=2, sticky=tk.W)
        self.leave_fields['days'] = ttk.Entry(row2_frame, width=10)
        self.leave_fields['days'].grid(row=0, column=3, padx=5, pady=2)
        
        ttk.Label(row2_frame, text="狀態*").grid(row=0, column=4, padx=5, pady=2, sticky=tk.W)
        self.leave_fields['status'] = ttk.Combobox(row2_frame, width=12, state="readonly",
                                                 values=["待審核", "已核准", "已拒絕"])
        self.leave_fields['status'].grid(row=0, column=5, padx=5, pady=2)
        
        # 第三行
        row3_frame = ttk.Frame(add_leave_frame)
        row3_frame.pack(fill=tk.X, pady=2)
        
        ttk.Label(row3_frame, text="請假事由*").pack(side=tk.LEFT, padx=5)
        self.leave_fields['reason'] = tk.Text(row3_frame, width=80, height=3)
        self.leave_fields['reason'].pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # 按鈕
        leave_button_frame = ttk.Frame(add_leave_frame)
        leave_button_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(leave_button_frame, text="➕ 申請請假", command=self.add_leave_request).pack(side=tk.LEFT, padx=5)
        ttk.Button(leave_button_frame, text="✏️ 編輯申請", command=self.edit_leave_request).pack(side=tk.LEFT, padx=5)
        ttk.Button(leave_button_frame, text="🗑️ 刪除申請", command=self.delete_leave_request).pack(side=tk.LEFT, padx=5)
        ttk.Button(leave_button_frame, text="✅ 核准", command=lambda: self.update_leave_status("已核准")).pack(side=tk.LEFT, padx=5)
        ttk.Button(leave_button_frame, text="❌ 拒絕", command=lambda: self.update_leave_status("已拒絕")).pack(side=tk.LEFT, padx=5)
        
    def create_overtime_management_tab(self, parent):
        """建立加班管理頁籤"""
        overtime_frame = ttk.Frame(parent)
        parent.add(overtime_frame, text="⏰ 加班管理")
        
        # 上方：篩選區域
        filter_frame = ttk.LabelFrame(overtime_frame, text="篩選條件", padding=5)
        filter_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # 員工篩選
        ttk.Label(filter_frame, text="員工：").grid(row=0, column=0, padx=5, pady=2)
        self.overtime_employee_var = tk.StringVar()
        self.overtime_employee_combo = ttk.Combobox(filter_frame, textvariable=self.overtime_employee_var,
                                                  width=20, state="readonly")
        self.overtime_employee_combo.grid(row=0, column=1, padx=5, pady=2)
        
        # 狀態篩選
        ttk.Label(filter_frame, text="狀態：").grid(row=0, column=2, padx=5, pady=2)
        self.overtime_status_var = tk.StringVar()
        overtime_status_combo = ttk.Combobox(filter_frame, textvariable=self.overtime_status_var,
                                           values=["全部", "待審核", "已核准", "已拒絕"], width=15, state="readonly")
        overtime_status_combo.grid(row=0, column=3, padx=5, pady=2)
        overtime_status_combo.set("全部")
        
        ttk.Button(filter_frame, text="🔍 篩選", command=self.filter_overtime_records).grid(row=0, column=4, padx=10, pady=2)
        ttk.Button(filter_frame, text="🔄 重新整理", command=self.refresh_overtime_records).grid(row=0, column=5, padx=5, pady=2)
        
        # 中間：加班記錄列表
        overtime_list_frame = ttk.LabelFrame(overtime_frame, text="加班記錄", padding=5)
        overtime_list_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        overtime_columns = ('員工編號', '員工姓名', '加班日期', '開始時間', '結束時間', '加班時數', '加班類型', '申請日期', '狀態', '備註')
        self.overtime_tree = ttk.Treeview(overtime_list_frame, columns=overtime_columns, show='headings', height=10)
        
        for col in overtime_columns:
            self.overtime_tree.heading(col, text=col)
            self.overtime_tree.column(col, width=85)
        
        overtime_scrollbar = ttk.Scrollbar(overtime_list_frame, orient=tk.VERTICAL, command=self.overtime_tree.yview)
        self.overtime_tree.configure(yscrollcommand=overtime_scrollbar.set)
        
        self.overtime_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        overtime_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 下方：新增加班申請
        add_overtime_frame = ttk.LabelFrame(overtime_frame, text="新增加班申請", padding=10)
        add_overtime_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.overtime_fields = {}
        
        # 第一行
        ot_row1_frame = ttk.Frame(add_overtime_frame)
        ot_row1_frame.pack(fill=tk.X, pady=2)
        
        ttk.Label(ot_row1_frame, text="員工*").grid(row=0, column=0, padx=5, pady=2, sticky=tk.W)
        self.overtime_emp_var = tk.StringVar()
        self.overtime_fields['employee'] = ttk.Combobox(ot_row1_frame, textvariable=self.overtime_emp_var, 
                                                      width=15, state="readonly")
        self.overtime_fields['employee'].grid(row=0, column=1, padx=5, pady=2)
        
        ttk.Label(ot_row1_frame, text="加班日期*").grid(row=0, column=2, padx=5, pady=2, sticky=tk.W)
        self.overtime_fields['overtime_date'] = ttk.Entry(ot_row1_frame, width=12)
        self.overtime_fields['overtime_date'].grid(row=0, column=3, padx=5, pady=2)
        self.overtime_fields['overtime_date'].insert(0, datetime.now().strftime("%Y-%m-%d"))
        
        ttk.Label(ot_row1_frame, text="開始時間*").grid(row=0, column=4, padx=5, pady=2, sticky=tk.W)
        self.overtime_fields['start_time'] = ttk.Entry(ot_row1_frame, width=10)
        self.overtime_fields['start_time'].grid(row=0, column=5, padx=5, pady=2)
        
        # 第二行
        ot_row2_frame = ttk.Frame(add_overtime_frame)
        ot_row2_frame.pack(fill=tk.X, pady=2)
        
        ttk.Label(ot_row2_frame, text="結束時間*").grid(row=0, column=0, padx=5, pady=2, sticky=tk.W)
        self.overtime_fields['end_time'] = ttk.Entry(ot_row2_frame, width=10)
        self.overtime_fields['end_time'].grid(row=0, column=1, padx=5, pady=2)
        
        ttk.Label(ot_row2_frame, text="加班時數*").grid(row=0, column=2, padx=5, pady=2, sticky=tk.W)
        self.overtime_fields['hours'] = ttk.Entry(ot_row2_frame, width=10)
        self.overtime_fields['hours'].grid(row=0, column=3, padx=5, pady=2)
        
        ttk.Label(ot_row2_frame, text="加班類型*").grid(row=0, column=4, padx=5, pady=2, sticky=tk.W)
        self.overtime_fields['overtime_type'] = ttk.Combobox(ot_row2_frame, width=15, state="readonly",
                                                           values=["平日加班", "假日加班", "國定假日加班"])
        self.overtime_fields['overtime_type'].grid(row=0, column=5, padx=5, pady=2)
        
        # 第三行
        ot_row3_frame = ttk.Frame(add_overtime_frame)
        ot_row3_frame.pack(fill=tk.X, pady=2)
        
        ttk.Label(ot_row3_frame, text="狀態*").grid(row=0, column=0, padx=5, pady=2, sticky=tk.W)
        self.overtime_fields['status'] = ttk.Combobox(ot_row3_frame, width=12, state="readonly",
                                                    values=["待審核", "已核准", "已拒絕"])
        self.overtime_fields['status'].grid(row=0, column=1, padx=5, pady=2)
        
        # 第四行
        ot_row4_frame = ttk.Frame(add_overtime_frame)
        ot_row4_frame.pack(fill=tk.X, pady=2)
        
        ttk.Label(ot_row4_frame, text="加班事由*").pack(side=tk.LEFT, padx=5)
        self.overtime_fields['reason'] = tk.Text(ot_row4_frame, width=80, height=3)
        self.overtime_fields['reason'].pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        # 按鈕
        overtime_button_frame = ttk.Frame(add_overtime_frame)
        overtime_button_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(overtime_button_frame, text="➕ 申請加班", command=self.add_overtime_request).pack(side=tk.LEFT, padx=5)
        ttk.Button(overtime_button_frame, text="✏️ 編輯申請", command=self.edit_overtime_request).pack(side=tk.LEFT, padx=5)
        ttk.Button(overtime_button_frame, text="🗑️ 刪除申請", command=self.delete_overtime_request).pack(side=tk.LEFT, padx=5)
        ttk.Button(overtime_button_frame, text="✅ 核准", command=lambda: self.update_overtime_status("已核准")).pack(side=tk.LEFT, padx=5)
        ttk.Button(overtime_button_frame, text="❌ 拒絕", command=lambda: self.update_overtime_status("已拒絕")).pack(side=tk.LEFT, padx=5)
    
    # === 工具方法 ===
    def get_widget_value(self, widget):
        """獲取控件的值"""
        if isinstance(widget, ttk.Entry) or isinstance(widget, ttk.Combobox):
            return widget.get()
        elif isinstance(widget, tk.Text):
            return widget.get("1.0", tk.END).strip()
        return ""
    
    def set_widget_value(self, widget, value):
        """設置控件的值"""
        if isinstance(widget, ttk.Entry) or isinstance(widget, ttk.Combobox):
            widget.delete(0, tk.END)
            widget.insert(0, str(value))
        elif isinstance(widget, tk.Text):
            widget.delete("1.0", tk.END)
            widget.insert("1.0", str(value))
    
    def validate_required_fields(self, fields_dict, field_configs):
        """驗證必填欄位"""
        missing_fields = []
        for field_key, config in field_configs.items():
            if config.get('required', False):
                value = self.get_widget_value(config['widget'])
                if not value.strip():
                    missing_fields.append(field_key)
        
        if missing_fields:
            messagebox.showerror("驗證錯誤", f"以下必填欄位不能為空：\n{', '.join(missing_fields)}")
            return False
        return True
    
    # === 員工管理相關方法 ===
    def new_employee(self):
        """新增員工"""
        self.current_employee_id = None
        self.clear_employee_form()
    
    def delete_employee(self):
        """刪除員工"""
        selection = self.employee_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "請先選擇要刪除的員工！")
            return
        
        item = self.employee_tree.selection()[0]
        employee_id = self.employee_tree.item(item)['values'][0]
        
        if messagebox.askyesno("確認", f"確定要刪除員工 {employee_id} 的所有資料嗎？此操作無法復原！"):
            if employee_id in self.employees_data:
                del self.employees_data[employee_id]
            self.refresh_employee_tree()
            self.refresh_employee_combos()
            self.clear_employee_form()
            messagebox.showinfo("成功", "員工資料已刪除！")
    
    def on_employee_select(self, event):
        """當選擇員工時"""
        selection = self.employee_tree.selection()
        if selection:
            item = selection[0]
            employee_id = self.employee_tree.item(item)['values'][0]
            self.current_employee_id = employee_id
            self.load_employee_data(employee_id)
    
    def load_employee_data(self, employee_id):
        """載入員工資料到表單"""
        if employee_id in self.employees_data:
            employee = self.employees_data[employee_id]
            basic_info = employee.get('basic_info', {})
            
            for field_key, config in self.basic_fields.items():
                value = basic_info.get(field_key, "")
                self.set_widget_value(config['widget'], value)
    
    def save_employee_info(self):
        """儲存員工資料"""
        if not self.validate_required_fields(self.basic_fields, self.basic_fields):
            return
        
        # 獲取員工編號
        employee_id = self.get_widget_value(self.basic_fields['employee_id']['widget'])
        
        if not employee_id:
            messagebox.showerror("錯誤", "員工編號不能為空！")
            return
        
        # 檢查是否重複（新增時）
        if self.current_employee_id != employee_id and employee_id in self.employees_data:
            messagebox.showerror("錯誤", "員工編號已存在！")
            return
        
        # 收集基本資料
        basic_data = {}
        for field_key, config in self.basic_fields.items():
            basic_data[field_key] = self.get_widget_value(config['widget'])
        
        # 初始化員工資料結構
        if employee_id not in self.employees_data:
            self.employees_data[employee_id] = {
                'basic_info': {},
                'performance_records': [],
                'leave_requests': [],
                'overtime_requests': []
            }
        
        # 如果是修改員工編號
        if self.current_employee_id and self.current_employee_id != employee_id:
            # 複製舊資料到新編號
            self.employees_data[employee_id] = self.employees_data[self.current_employee_id].copy()
            # 刪除舊資料
            del self.employees_data[self.current_employee_id]
        
        # 儲存基本資料
        self.employees_data[employee_id]['basic_info'] = basic_data
        self.current_employee_id = employee_id
        
        # 刷新顯示
        self.refresh_employee_tree()
        self.refresh_employee_combos()
        
        messagebox.showinfo("成功", "員工資料已儲存！")
    
    def clear_employee_form(self):
        """清空員工表單"""
        for config in self.basic_fields.values():
            self.set_widget_value(config['widget'], "")
    
    def refresh_employee_tree(self):
        """刷新員工列表"""
        # 清空現有項目
        for item in self.employee_tree.get_children():
            self.employee_tree.delete(item)
        
        # 重新載入資料
        for employee_id, employee_data in self.employees_data.items():
            basic_info = employee_data.get('basic_info', {})
            values = (
                employee_id,
                basic_info.get('name', ''),
                basic_info.get('department', ''),
                basic_info.get('position', ''),
                basic_info.get('hire_date', '')
            )
            self.employee_tree.insert('', 'end', values=values)
    
    def refresh_employee_combos(self):
        """刷新所有員工下拉選單"""
        employee_list = [f"{emp_id} - {data['basic_info'].get('name', '')}" 
                        for emp_id, data in self.employees_data.items()]
        
        # 更新考績管理的員工選單
        self.perf_employee_combo['values'] = employee_list
        
        # 更新請假管理的員工選單
        self.leave_employee_combo['values'] = employee_list
        self.leave_fields['employee']['values'] = employee_list
        
        # 更新加班管理的員工選單
        self.overtime_employee_combo['values'] = employee_list
        self.overtime_fields['employee']['values'] = employee_list
    
    # === 考績管理相關方法 ===
    def on_perf_employee_select(self, event):
        """當選擇考績管理的員工時"""
        selected = self.perf_employee_var.get()
        if selected:
            employee_id = selected.split(' - ')[0]
            self.refresh_performance_tree(employee_id)
    
    def refresh_performance_tree(self, employee_id=None):
        """刷新考績記錄表格"""
        # 清空現有項目
        for item in self.perf_tree.get_children():
            self.perf_tree.delete(item)
        
        if not employee_id:
            return
        
        # 重新載入資料
        if employee_id in self.employees_data:
            records = self.employees_data[employee_id].get('performance_records', [])
            for record in records:
                values = (
                    record.get('year', ''),
                    record.get('first_half', ''),
                    record.get('second_half', ''),
                    record.get('annual_rating', ''),
                    record.get('remarks', '')
                )
                self.perf_tree.insert('', 'end', values=values)
    
    def add_performance(self):
        """新增考績記錄"""
        selected = self.perf_employee_var.get()
        if not selected:
            messagebox.showerror("錯誤", "請先選擇員工！")
            return
        
        employee_id = selected.split(' - ')[0]
        
        # 驗證必填欄位
        required_fields = ['year', 'annual_rating']
        for field in required_fields:
            if not self.get_widget_value(self.perf_fields[field]).strip():
                messagebox.showerror("錯誤", f"{field} 為必填欄位！")
                return
        
        # 獲取資料
        perf_data = {}
        for field_key, widget in self.perf_fields.items():
            perf_data[field_key] = self.get_widget_value(widget)
        
        # 加入到記錄中
        if employee_id in self.employees_data:
            self.employees_data[employee_id]['performance_records'].append(perf_data)
            
            # 更新表格顯示
            self.refresh_performance_tree(employee_id)
            
            # 清空輸入欄位
            for widget in self.perf_fields.values():
                self.set_widget_value(widget, "")
            
            messagebox.showinfo("成功", "考績記錄已新增！")
    
    def edit_performance(self):
        """編輯考績記錄"""
        selection = self.perf_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "請先選擇要編輯的考績記錄！")
            return
        
        selected_emp = self.perf_employee_var.get()
        if not selected_emp:
            return
        
        employee_id = selected_emp.split(' - ')[0]
        item = self.perf_tree.selection()[0]
        index = self.perf_tree.index(item)
        
        # 將資料填入輸入欄位
        record = self.employees_data[employee_id]['performance_records'][index]
        for field_key, widget in self.perf_fields.items():
            if field_key in record:
                self.set_widget_value(widget, record[field_key])
        
        # 刪除舊記錄
        del self.employees_data[employee_id]['performance_records'][index]
        self.refresh_performance_tree(employee_id)
    
    def delete_performance(self):
        """刪除考績記錄"""
        selection = self.perf_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "請先選擇要刪除的考績記錄！")
            return
        
        if messagebox.askyesno("確認", "確定要刪除選中的考績記錄嗎？"):
            selected_emp = self.perf_employee_var.get()
            if selected_emp:
                employee_id = selected_emp.split(' - ')[0]
                item = self.perf_tree.selection()[0]
                index = self.perf_tree.index(item)
                del self.employees_data[employee_id]['performance_records'][index]
                self.refresh_performance_tree(employee_id)
                messagebox.showinfo("成功", "考績記錄已刪除！")
    
    # === 請假管理相關方法 ===
    def refresh_leave_records(self):
        """刷新請假記錄"""
        self.filter_leave_records()
    
    def filter_leave_records(self):
        """篩選請假記錄"""
        # 清空現有項目
        for item in self.leave_tree.get_children():
            self.leave_tree.delete(item)
        
        # 獲取篩選條件
        selected_emp = self.leave_employee_var.get()
        selected_status = self.leave_status_var.get()
        
        # 載入資料
        for employee_id, employee_data in self.employees_data.items():
            # 員工篩選
            if selected_emp and not selected_emp.startswith("全部"):
                if selected_emp and selected_emp.split(' - ')[0] != employee_id:
                    continue
            
            basic_info = employee_data.get('basic_info', {})
            employee_name = basic_info.get('name', '')
            
            records = employee_data.get('leave_requests', [])
            for record in records:
                # 狀態篩選
                if selected_status and selected_status != "全部":
                    if record.get('status', '') != selected_status:
                        continue
                
                values = (
                    employee_id,
                    employee_name,
                    record.get('leave_type', ''),
                    record.get('start_date', ''),
                    record.get('end_date', ''),
                    record.get('days', ''),
                    record.get('apply_date', ''),
                    record.get('status', ''),
                    record.get('reason', '')
                )
                self.leave_tree.insert('', 'end', values=values)
    
    def add_leave_request(self):
        """新增請假申請"""
        selected_emp = self.leave_emp_var.get()
        if not selected_emp:
            messagebox.showerror("錯誤", "請先選擇員工！")
            return
        
        employee_id = selected_emp.split(' - ')[0]
        
        # 驗證必填欄位
        required_fields = ['leave_type', 'start_date', 'end_date', 'days', 'status', 'reason']
        for field in required_fields:
            if field == 'reason':
                value = self.get_widget_value(self.leave_fields[field])
            else:
                value = self.get_widget_value(self.leave_fields[field])
            if not value.strip():
                messagebox.showerror("錯誤", f"{field} 為必填欄位！")
                return
        
        # 獲取資料
        leave_data = {
            'apply_date': datetime.now().strftime("%Y-%m-%d")
        }
        for field_key, widget in self.leave_fields.items():
            if field_key != 'employee':
                leave_data[field_key] = self.get_widget_value(widget)
        
        # 加入到記錄中
        if employee_id in self.employees_data:
            if 'leave_requests' not in self.employees_data[employee_id]:
                self.employees_data[employee_id]['leave_requests'] = []
            self.employees_data[employee_id]['leave_requests'].append(leave_data)
            
            # 更新表格顯示
            self.refresh_leave_records()
            
            # 清空輸入欄位
            for field_key, widget in self.leave_fields.items():
                if field_key != 'employee':
                    self.set_widget_value(widget, "")
            
            messagebox.showinfo("成功", "請假申請已提交！")
    
    def edit_leave_request(self):
        """編輯請假申請"""
        selection = self.leave_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "請先選擇要編輯的請假申請！")
            return
        
        item = self.leave_tree.selection()[0]
        values = self.leave_tree.item(item)['values']
        employee_id = values[0]
        
        # 找到對應的記錄
        if employee_id in self.employees_data:
            records = self.employees_data[employee_id].get('leave_requests', [])
            # 這裡簡化處理，實際應該用更精確的方法找到記錄
            for i, record in enumerate(records):
                if (record.get('leave_type') == values[2] and 
                    record.get('start_date') == values[3]):
                    
                    # 設置員工選擇
                    for emp_option in self.leave_fields['employee']['values']:
                        if emp_option.startswith(employee_id):
                            self.leave_emp_var.set(emp_option)
                            break
                    
                    # 填充其他欄位
                    field_mapping = {
                        'leave_type': values[2],
                        'start_date': values[3],
                        'end_date': values[4],
                        'days': values[5],
                        'status': values[7],
                        'reason': values[8]
                    }
                    
                    for field_key, value in field_mapping.items():
                        if field_key in self.leave_fields:
                            self.set_widget_value(self.leave_fields[field_key], value)
                    
                    # 刪除舊記錄
                    del records[i]
                    self.refresh_leave_records()
                    break
    
    def delete_leave_request(self):
        """刪除請假申請"""
        selection = self.leave_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "請先選擇要刪除的請假申請！")
            return
        
        if messagebox.askyesno("確認", "確定要刪除選中的請假申請嗎？"):
            item = self.leave_tree.selection()[0]
            values = self.leave_tree.item(item)['values']
            employee_id = values[0]
            
            # 找到並刪除記錄
            if employee_id in self.employees_data:
                records = self.employees_data[employee_id].get('leave_requests', [])
                for i, record in enumerate(records):
                    if (record.get('leave_type') == values[2] and 
                        record.get('start_date') == values[3]):
                        del records[i]
                        break
                
                self.refresh_leave_records()
                messagebox.showinfo("成功", "請假申請已刪除！")
    
    def update_leave_status(self, new_status):
        """更新請假狀態"""
        selection = self.leave_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "請先選擇要更新的請假申請！")
            return
        
        item = self.leave_tree.selection()[0]
        values = self.leave_tree.item(item)['values']
        employee_id = values[0]
        
        # 找到並更新記錄
        if employee_id in self.employees_data:
            records = self.employees_data[employee_id].get('leave_requests', [])
            for record in records:
                if (record.get('leave_type') == values[2] and 
                    record.get('start_date') == values[3]):
                    record['status'] = new_status
                    break
            
            self.refresh_leave_records()
            messagebox.showinfo("成功", f"請假狀態已更新為：{new_status}")
    
    # === 加班管理相關方法 ===
    def refresh_overtime_records(self):
        """刷新加班記錄"""
        self.filter_overtime_records()
    
    def filter_overtime_records(self):
        """篩選加班記錄"""
        # 清空現有項目
        for item in self.overtime_tree.get_children():
            self.overtime_tree.delete(item)
        
        # 獲取篩選條件
        selected_emp = self.overtime_employee_var.get()
        selected_status = self.overtime_status_var.get()
        
        # 載入資料
        for employee_id, employee_data in self.employees_data.items():
            # 員工篩選
            if selected_emp and not selected_emp.startswith("全部"):
                if selected_emp and selected_emp.split(' - ')[0] != employee_id:
                    continue
            
            basic_info = employee_data.get('basic_info', {})
            employee_name = basic_info.get('name', '')
            
            records = employee_data.get('overtime_requests', [])
            for record in records:
                # 狀態篩選
                if selected_status and selected_status != "全部":
                    if record.get('status', '') != selected_status:
                        continue
                
                values = (
                    employee_id,
                    employee_name,
                    record.get('overtime_date', ''),
                    record.get('start_time', ''),
                    record.get('end_time', ''),
                    record.get('hours', ''),
                    record.get('overtime_type', ''),
                    record.get('apply_date', ''),
                    record.get('status', ''),
                    record.get('reason', '')
                )
                self.overtime_tree.insert('', 'end', values=values)
    
    def add_overtime_request(self):
        """新增加班申請"""
        selected_emp = self.overtime_emp_var.get()
        if not selected_emp:
            messagebox.showerror("錯誤", "請先選擇員工！")
            return
        
        employee_id = selected_emp.split(' - ')[0]
        
        # 驗證必填欄位
        required_fields = ['overtime_date', 'start_time', 'end_time', 'hours', 'overtime_type', 'status', 'reason']
        for field in required_fields:
            value = self.get_widget_value(self.overtime_fields[field])
            if not value.strip():
                messagebox.showerror("錯誤", f"{field} 為必填欄位！")
                return
        
        # 獲取資料
        overtime_data = {
            'apply_date': datetime.now().strftime("%Y-%m-%d")
        }
        for field_key, widget in self.overtime_fields.items():
            if field_key != 'employee':
                overtime_data[field_key] = self.get_widget_value(widget)
        
        # 加入到記錄中
        if employee_id in self.employees_data:
            if 'overtime_requests' not in self.employees_data[employee_id]:
                self.employees_data[employee_id]['overtime_requests'] = []
            self.employees_data[employee_id]['overtime_requests'].append(overtime_data)
            
            # 更新表格顯示
            self.refresh_overtime_records()
            
            # 清空輸入欄位
            for field_key, widget in self.overtime_fields.items():
                if field_key != 'employee':
                    self.set_widget_value(widget, "")
            
            # 重置預設值
            self.set_widget_value(self.overtime_fields['overtime_date'], datetime.now().strftime("%Y-%m-%d"))
            
            messagebox.showinfo("成功", "加班申請已提交！")
    
    def edit_overtime_request(self):
        """編輯加班申請"""
        selection = self.overtime_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "請先選擇要編輯的加班申請！")
            return
        
        item = self.overtime_tree.selection()[0]
        values = self.overtime_tree.item(item)['values']
        employee_id = values[0]
        
        # 找到對應的記錄
        if employee_id in self.employees_data:
            records = self.employees_data[employee_id].get('overtime_requests', [])
            for i, record in enumerate(records):
                if (record.get('overtime_date') == values[2] and 
                    record.get('start_time') == values[3]):
                    
                    # 設置員工選擇
                    for emp_option in self.overtime_fields['employee']['values']:
                        if emp_option.startswith(employee_id):
                            self.overtime_emp_var.set(emp_option)
                            break

                    # 填充其他欄位
                    field_mapping = {
                        'overtime_date': values[2],
                        'start_time': values[3],
                        'end_time': values[4],
                        'hours': values[5],
                        'overtime_type': values[6],
                        'status': values[8],
                        'reason': values[9]
                    }
                    for field_key, value in field_mapping.items():
                        if field_key in self.overtime_fields:
                            self.set_widget_value(self.overtime_fields[field_key], value)

                    # 刪除舊記錄
                    del records[i]
                    self.refresh_overtime_records()
                    break

    def delete_overtime_request(self):
        """刪除加班申請"""
        selection = self.overtime_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "請先選擇要刪除的加班申請！")
            return

        if messagebox.askyesno("確認", "確定要刪除選中的加班申請嗎？"):
            item = self.overtime_tree.selection()[0]
            values = self.overtime_tree.item(item)['values']
            employee_id = values[0]

            # 找到並刪除記錄
            if employee_id in self.employees_data:
                records = self.employees_data[employee_id].get('overtime_requests', [])
                for i, record in enumerate(records):
                    if (record.get('overtime_date') == values[2] and
                        record.get('start_time') == values[3]):
                        del records[i]
                        break

                self.refresh_overtime_records()
                messagebox.showinfo("成功", "加班申請已刪除！")

    def update_overtime_status(self, new_status):
        """更新加班狀態"""
        selection = self.overtime_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "請先選擇要更新的加班申請！")
            return

        item = self.overtime_tree.selection()[0]
        values = self.overtime_tree.item(item)['values']
        employee_id = values[0]

        # 找到並更新記錄
        if employee_id in self.employees_data:
            records = self.employees_data[employee_id].get('overtime_requests', [])
            for record in records:
                if (record.get('overtime_date') == values[2] and
                    record.get('start_time') == values[3]):
                    record['status'] = new_status
                    break

            self.refresh_overtime_records()
            messagebox.showinfo("成功", f"加班狀態已更新為：{new_status}")

    # === 檔案與資料處理 ===
    def clear_all_data(self):
        """清空所有資料"""
        if messagebox.askyesno("確認", "確定要清空所有員工資料與申請紀錄嗎？此操作無法復原！"):
            self.employees_data.clear()
            self.current_employee_id = None
            self.refresh_employee_tree()
            self.refresh_employee_combos()
            self.refresh_performance_tree()
            self.refresh_leave_records()
            self.refresh_overtime_records()
            self.clear_employee_form()
            messagebox.showinfo("成功", "所有資料已清空！")

    def save_data(self):
        """儲存所有資料至本地 JSON 檔案"""
        file_path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON檔案", "*.json")],
            title="儲存資料"
        )
        if not file_path:
            return

        try:
            with open(file_path, "w", encoding="utf-8") as f:
                json.dump(self.employees_data, f, ensure_ascii=False, indent=2)
            messagebox.showinfo("成功", "資料已儲存至檔案！")
        except Exception as e:
            messagebox.showerror("錯誤", f"儲存檔案失敗：{e}")

    def load_data(self):
        """從本地 JSON 檔案載入資料"""
        file_path = filedialog.askopenfilename(
            defaultextension=".json",
            filetypes=[("JSON檔案", "*.json")],
            title="載入資料"
        )
        if not file_path:
            return

        try:
            with open(file_path, "r", encoding="utf-8") as f:
                self.employees_data = json.load(f)
            self.refresh_employee_tree()
            self.refresh_employee_combos()
            self.refresh_performance_tree()
            self.refresh_leave_records()
            self.refresh_overtime_records()
            self.clear_employee_form()
            messagebox.showinfo("成功", "資料已載入！")
        except Exception as e:
            messagebox.showerror("錯誤", f"載入檔案失敗：{e}")

    def import_excel(self):
        """從 Excel 匯入員工資料（僅支援格式示範）"""
        file_path = filedialog.askopenfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel檔案", "*.xlsx")],
            title="匯入Excel"
        )
        if not file_path:
            return

        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            # 假設第一列是欄位名稱
            rows = list(ws.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                messagebox.showerror("錯誤", "Excel檔案內容不足")
                return

            header = rows[0]
            for row in rows[1:]:
                employee_data = dict(zip(header, row))
                emp_id = str(employee_data.get('employee_id', '')).strip()
                if emp_id:
                    self.employees_data[emp_id] = self.employees_data.get(emp_id, {
                        'basic_info': {}, 'performance_records': [], 'leave_requests': [], 'overtime_requests': []
                    })
                    # 只匯入基本資料
                    self.employees_data[emp_id]['basic_info'] = employee_data

            self.refresh_employee_tree()
            self.refresh_employee_combos()
            messagebox.showinfo("成功", "Excel匯入完成（僅匯入基本資料）！")
        except Exception as e:
            messagebox.showerror("錯誤", f"Excel匯入失敗：{e}")

    def export_excel(self):
        """匯出所有員工資料為分頁、多樣式的Excel"""
        file_path = filedialog.asksaveasfilename(
            title="儲存Excel檔案",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            # 預設sheet刪除
            wb.remove(wb.active)

            # 1. 員工基本資料
            self.export_basic_info_sheet(wb)
            # 2. 員工考績
            self.export_performance_sheet(wb)
            # 3. 員工請假
            self.export_leave_sheet(wb)
            # 4. 員工加班
            self.export_overtime_sheet(wb)

            wb.save(file_path)
            messagebox.showinfo("成功", f"資料已匯出到：{file_path}")
        except Exception as e:
            messagebox.showerror("錯誤", f"匯出失敗：{e}")

    def export_basic_info_sheet(self, wb):
        ws = wb.create_sheet("員工基本資料")
        # 決定所有欄位
        headers = [
            "員工編號", "姓名", "身分證字號", "性別", "出生日期", "聯絡電話", "電子郵件",
            "緊急聯絡人", "緊急聯絡人電話", "戶籍地址", "通訊地址",
            "部門", "職位", "職級", "到職日期", "直屬主管", "工作地點", "僱用類型", "薪資等級"
        ]
        self.write_styled_sheet(ws, headers, [
            [
                emp_id,
                info.get("name", ""),
                info.get("id_number", ""),
                info.get("gender", ""),
                info.get("birth_date", ""),
                info.get("phone", ""),
                info.get("email", ""),
                info.get("emergency_contact", ""),
                info.get("emergency_phone", ""),
                info.get("address", ""),
                info.get("mailing_address", ""),
                info.get("department", ""),
                info.get("position", ""),
                info.get("job_level", ""),
                info.get("hire_date", ""),
                info.get("supervisor", ""),
                info.get("work_location", ""),
                info.get("employment_type", ""),
                info.get("salary_grade", ""),
            ]
            for emp_id, data in self.employees_data.items()
            for info in [data.get('basic_info', {})]
        ])

    def export_performance_sheet(self, wb):
        ws = wb.create_sheet("考績管理")
        headers = ["員工編號", "姓名", "年度", "上半年考績", "下半年考績", "年度總評", "備註"]
        data = []
        for emp_id, emp_data in self.employees_data.items():
            name = emp_data.get("basic_info", {}).get("name", "")
            for perf in emp_data.get("performance_records", []):
                data.append([
                    emp_id,
                    name,
                    perf.get("year", ""),
                    perf.get("first_half", ""),
                    perf.get("second_half", ""),
                    perf.get("annual_rating", ""),
                    perf.get("remarks", ""),
                ])
        self.write_styled_sheet(ws, headers, data)

    def export_leave_sheet(self, wb):
        ws = wb.create_sheet("請假管理")
        headers = ["員工編號", "姓名", "請假類型", "開始日期", "結束日期", "請假天數", "申請日期", "狀態", "請假事由"]
        data = []
        for emp_id, emp_data in self.employees_data.items():
            name = emp_data.get("basic_info", {}).get("name", "")
            for leave in emp_data.get("leave_requests", []):
                data.append([
                    emp_id,
                    name,
                    leave.get("leave_type", ""),
                    leave.get("start_date", ""),
                    leave.get("end_date", ""),
                    leave.get("days", ""),
                    leave.get("apply_date", ""),
                    leave.get("status", ""),
                    leave.get("reason", ""),
                ])
        self.write_styled_sheet(ws, headers, data)

    def export_overtime_sheet(self, wb):
        ws = wb.create_sheet("加班管理")
        headers = ["員工編號", "姓名", "加班日期", "開始時間", "結束時間", "加班時數", "加班類型", "申請日期", "狀態", "加班事由"]
        data = []
        for emp_id, emp_data in self.employees_data.items():
            name = emp_data.get("basic_info", {}).get("name", "")
            for ot in emp_data.get("overtime_requests", []):
                data.append([
                    emp_id,
                    name,
                    ot.get("overtime_date", ""),
                    ot.get("start_time", ""),
                    ot.get("end_time", ""),
                    ot.get("hours", ""),
                    ot.get("overtime_type", ""),
                    ot.get("apply_date", ""),
                    ot.get("status", ""),
                    ot.get("reason", ""),
                ])
        self.write_styled_sheet(ws, headers, data)

    def write_styled_sheet(self, ws, headers, data_rows):
        """寫入標題和資料, 標題有顏色，凍結首列，加上filter，自動欄寬"""
        # 標題樣式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))

        # 標題列
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # 資料列
        for row_idx, row_data in enumerate(data_rows, 2):
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border

        # filter
        ws.auto_filter.ref = ws.dimensions

        # 凍結首列
        ws.freeze_panes = "A2"

        # 自動欄寬
        for col in ws.columns:
            max_length = max([len(str(cell.value)) if cell.value else 0 for cell in col] + [len(str(col[0].value))])
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

# 主程式執行
if __name__ == '__main__':
    root = tk.Tk()
    app = EmployeeFormSystem(root)
    root.mainloop()
